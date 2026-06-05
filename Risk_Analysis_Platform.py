"""
RISK ANALYSIS PLATFORM - ENTERPRISE EDITION
Project Risk Management System | PMI-RMP Compliant | Monte Carlo Simulation
Features: Primavera P6 Integration | AI Risk Prediction | Real-Time Weather
          EMV Analysis | Correlation Modeling | Power BI Export
"""

import streamlit as st
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
import io
import zipfile
import json
import requests
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import warnings
import xml.etree.ElementTree as ET
from sklearn.ensemble import RandomForestRegressor
from sklearn.preprocessing import StandardScaler
from scipy.linalg import cholesky, eigh
from scipy.stats import norm, qmc
import time
import random
import threading
import hashlib
from functools import wraps
st.cache_data.clear()

warnings.filterwarnings('ignore')

# ============================================================================
# API RATE LIMITING
# ============================================================================
class RateLimiter:
    def __init__(self, calls_per_second=1):
        self.calls_per_second = calls_per_second
        self.last_call_time = 0
        self.lock = threading.Lock()
    
    def wait_if_needed(self):
        with self.lock:
            current_time = time.time()
            time_since_last = current_time - self.last_call_time
            if time_since_last < (1.0 / self.calls_per_second):
                time.sleep((1.0 / self.calls_per_second) - time_since_last)
            self.last_call_time = time.time()

geocode_limiter = RateLimiter(calls_per_second=0.5)

# ============================================================================
# STREAMLIT PAGE CONFIGURATION
# ============================================================================
st.set_page_config(
    page_title="Risk Analysis Platform",
    page_icon="⚙️",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ============================================================================
# GLOBAL CHART CONFIGURATION
# ============================================================================
chart_config = {
    'scrollZoom': False,   # ← CHANGE THIS FROM True TO False
    'displayModeBar': True,
    'responsive': True,
    'modeBarButtonsToRemove': ['lasso2d', 'select2d'],
    'modeBarButtonsToAdd': ['zoomIn', 'zoomOut', 'resetGeo', 'toImage'],
    'displaylogo': False,
    'staticPlot': False,
    'doubleClick': 'reset'
}

# ============================================================================
# CUSTOM CSS - DARK THEME
# ============================================================================
st.markdown("""
<style>
    /* VIEWPORT FIX - PREVENTS ZOOMING ON MOBILE */
    @viewport {
        width: device-width;
        zoom: 1.0;
    }
    
    /* Force mobile viewport */
    @-ms-viewport { width: device-width; }
    
    .stApp { background-color: #0f172a; }
    section[data-testid="stSidebar"] { width: 320px !important; }
    .stNumberInput input { text-align: left; }
    .dataframe { text-align: center !important; }
    .dataframe th { text-align: center !important; }
    .dataframe td { text-align: center !important; }
    
    /* CARDS */
    .card {
        background-color: #1e293b;
        border-radius: 12px;
        padding: 1.5rem;
        margin-bottom: 1.5rem;
        border: 1px solid #334155;
        box-shadow: 0 4px 6px -1px rgba(0,0,0,0.2);
    }
    .card-title {
        font-size: 1.125rem;
        font-weight: 600;
        color: #f1f5f9;
        margin-bottom: 1rem;
        padding-bottom: 0.5rem;
        border-bottom: 2px solid #334155;
    }
    .main-title {
        font-size: 2.5rem;
        font-weight: 700;
        color: #f1f5f9;
        text-align: center;
        letter-spacing: -0.02em;
        margin-bottom: 0.25rem;
    }
    .main-subtitle {
        font-size: 0.875rem;
        color: #94a3b8;
        text-align: center;
        margin-bottom: 2rem;
    }
    .metric-card {
        background-color: #0f172a;
        border-radius: 8px;
        padding: 1rem;
        text-align: center;
        border: 1px solid #334155;
    }
    .metric-card-label {
        font-size: 0.75rem;
        color: #94a3b8;
        text-transform: uppercase;
        font-weight: 500;
    }
    .metric-card-value {
        font-size: 1.5rem;
        font-weight: 700;
        color: #f1f5f9;
        margin-top: 0.5rem;
    }
    .stButton button {
        background-color: #3b82f6;
        color: white;
        border-radius: 6px;
        border: none;
        padding: 0.5rem 1rem;
        font-weight: 500;
        width: 100%;
    }
    .stButton button:hover { background-color: #2563eb; }
    .footer {
        text-align: center;
        padding: 1.5rem;
        margin-top: 2rem;
        border-top: 1px solid #334155;
        color: #64748b;
        font-size: 0.75rem;
    }
    
    /* ============================================ */
    /* MOBILE RESPONSIVE FIX - ALL ELEMENTS        */
    /* ============================================ */
    @media only screen and (max-width: 768px) {
        /* Prevent zoom on input focus (iOS) */
        input, textarea, select {
            font-size: 16px !important;
        }
        
        /* ===== TABLES ===== */
        .dataframe, .stDataFrame {
            width: 100% !important;
            overflow-x: auto !important;
            display: block !important;
            font-size: 11px !important;
        }
        
        .dataframe table, .stDataFrame table {
            width: 100% !important;
            min-width: 500px !important;
            font-size: 10px !important;
        }
        
        .dataframe th, .stDataFrame th,
        .dataframe td, .stDataFrame td {
            padding: 4px 5px !important;
            white-space: nowrap !important;
        }
        
        /* ===== PLOTLY CHARTS ===== */
        .js-plotly-plot, .plotly {
            width: 100% !important;
            max-width: 100% !important;
            overflow-x: auto !important;
        }
        
        .plotly .main-svg {
            width: 100% !important;
            max-width: 100% !important;
            height: auto !important;
        }
        
        /* Force all SVG charts to be responsive */
        svg:not(:root) {
            width: 100% !important;
            height: auto !important;
            max-width: 100% !important;
        }
        
        /* Make chart containers scrollable if needed */
        .stPlotlyChart {
            width: 100% !important;
            overflow-x: auto !important;
        }
        
        /* ===== METRICS ===== */
        .metric-card-value {
            font-size: 0.9rem !important;
        }
        
        /* ===== CARDS ===== */
        .card {
            padding: 0.75rem !important;
            border-radius: 8px !important;
        }
        
        .card-title {
            font-size: 0.95rem !important;
        }
        
        /* ===== HEADERS ===== */
        .main-title {
            font-size: 1.4rem !important;
        }
        
        .main-subtitle {
            font-size: 0.65rem !important;
            margin-bottom: 1rem !important;
        }
        
        /* ===== BUTTONS ===== */
        .stButton button {
            padding: 0.5rem 0.4rem !important;
            font-size: 0.75rem !important;
        }
        
        /* ===== FOOTER ===== */
        .footer {
            font-size: 0.55rem !important;
            padding: 0.6rem !important;
        }
        
        /* ===== SIDEBAR ===== */
        section[data-testid="stSidebar"] {
            width: 280px !important;
        }
        
        /* ===== EXPANDERS ===== */
        .streamlit-expanderHeader {
            font-size: 0.85rem !important;
        }
        
        /* ===== NUMBER INPUTS ===== */
        .stNumberInput input {
            font-size: 14px !important;
        }
        
        /* ===== DATA EDITOR ===== */
        .stDataEditor {
            font-size: 10px !important;
        }
        
        /* Force all containers to respect mobile width */
        .row-widget, .stHorizontalBlock {
            flex-wrap: wrap !important;
        }
        
        .stColumn {
            min-width: 100% !important;
            flex: 1 1 100% !important;
        }
    }
</style>

<script>
    /* JavaScript to force Plotly charts to be responsive on mobile */
    window.addEventListener('resize', function() {
        var plotlyElements = document.querySelectorAll('.js-plotly-plot');
        plotlyElements.forEach(function(el) {
            if (el && el.layout) {
                Plotly.relayout(el, {width: el.parentElement.clientWidth});
            }
        });
    });
    document.addEventListener('DOMContentLoaded', function() {
        var plotlyElements = document.querySelectorAll('.js-plotly-plot');
        plotlyElements.forEach(function(el) {
            if (el && el.parentElement) {
                Plotly.relayout(el, {width: el.parentElement.clientWidth});
            }
        });
    });
</script>
""", unsafe_allow_html=True)

# ============================================================================
# HEADER
# ============================================================================
st.markdown('<div class="main-title">RISK ANALYSIS PLATFORM</div>', unsafe_allow_html=True)
st.markdown('<div class="main-subtitle">Enterprise Risk Management System | PMI-RMP Compliant | Monte Carlo Simulation | AI-Powered</div>', unsafe_allow_html=True)

# ============================================================================
# INPUT VALIDATION FUNCTIONS
# ============================================================================
def validate_risk_input(risk_df):
    """Validate risk dataframe inputs"""
    errors = []
    
    if not risk_df['Probability (0-1)'].between(0, 1).all():
        errors.append("Probabilities must be between 0 and 1")
    
    if not risk_df['Impact (0-1)'].between(0, 1).all():
        errors.append("Impacts must be between 0 and 1")
    
    for idx, row in risk_df.iterrows():
        opt = row['Optimistic (0-1)']
        ml = row['Most Likely (0-1)']
        pess = row['Pessimistic (0-1)']
        
        if not (opt <= ml <= pess):
            errors.append(f"Risk '{row['Risk Name']}': Must satisfy Optimistic ≤ Most Likely ≤ Pessimistic")
        
        if not (0 <= opt <= 1 and 0 <= ml <= 1 and 0 <= pess <= 1):
            errors.append(f"Risk '{row['Risk Name']}': All distribution parameters must be between 0 and 1")
    
    if 'Correlation Strength' in risk_df.columns:
        if not risk_df['Correlation Strength'].between(0, 1).all():
            errors.append("Correlation strengths must be between 0 and 1")
    
    if risk_df['Risk Name'].duplicated().any():
        errors.append("Duplicate risk names found")
    
    if (risk_df['Risk Name'] == '').any():
        errors.append("Empty risk names not allowed")
    
    return errors

def validate_project_parameters(duration, direct_cost, indirect_rate):
    """Validate project parameters"""
    errors = []
    if duration <= 0:
        errors.append("Project duration must be positive")
    if direct_cost < 0:
        errors.append("Direct cost cannot be negative")
    if indirect_rate < 0:
        errors.append("Indirect rate cannot be negative")
    return errors

# ============================================================================
# CORRELATION MATRIX UTILITIES
# ============================================================================
def make_positive_definite(matrix, epsilon=1e-8):
    """Ensure matrix is positive definite using eigenvalue correction"""
    try:
        cholesky(matrix, lower=True)
        return matrix
    except:
        eigenvalues, eigenvectors = eigh(matrix)
        eigenvalues = np.maximum(eigenvalues, epsilon)
        corrected_matrix = eigenvectors @ np.diag(eigenvalues) @ eigenvectors.T
        d = np.sqrt(np.diag(corrected_matrix))
        corrected_matrix = corrected_matrix / np.outer(d, d)
        return corrected_matrix

def advanced_correlation_model(risk_df, n_iterations):
    """Generate correlated uniform samples using Cholesky decomposition"""
    n_risks = len(risk_df)
    corr_matrix = np.eye(n_risks)
    
    for i, row in risk_df.iterrows():
        correlated_with = row.get('Correlated With', '')
        if correlated_with and correlated_with != '':
            matching = risk_df[risk_df['Risk Name'] == correlated_with]
            if len(matching) > 0:
                j = matching.index[0]
                corr_strength = min(1.0, max(0.0, row.get('Correlation Strength', 0.0)))
                corr_matrix[i, j] = corr_strength
                corr_matrix[j, i] = corr_strength
    
    corr_matrix = make_positive_definite(corr_matrix)
    
    try:
        L = cholesky(corr_matrix, lower=True)
        uncorrelated = np.random.randn(n_iterations, n_risks)
        correlated = uncorrelated @ L.T
        correlated_uniform = norm.cdf(correlated)
        return correlated_uniform
    except:
        return np.random.rand(n_iterations, n_risks)

# ============================================================================
# EXPECTED MONETARY VALUE (EMV) CALCULATIONS
# ============================================================================
def calculate_correct_emv(risk_row, direct_cost, indirect_rate, project_duration):
    """Calculate EMV considering both cost and schedule impacts"""
    prob = risk_row['Probability (0-1)']
    impact_percentage = risk_row['Impact (0-1)']
    cost_impact = direct_cost * impact_percentage
    time_impact_days = project_duration * impact_percentage
    indirect_impact = time_impact_days * indirect_rate
    total_impact = cost_impact + indirect_impact
    emv = prob * total_impact
    return emv, total_impact

def get_response_effectiveness(response_strategy, risk_emv):
    """Calculate risk response effectiveness based on strategy type"""
    effectiveness = {
        'Avoid': {'probability_reduction': 0.9, 'impact_reduction': 0.95, 'cost_multiplier': 0.15},
        'Mitigate': {'probability_reduction': 0.5, 'impact_reduction': 0.6, 'cost_multiplier': 0.08},
        'Transfer': {'probability_reduction': 0.1, 'impact_reduction': 0.8, 'cost_multiplier': 0.12},
        'Accept': {'probability_reduction': 0.0, 'impact_reduction': 0.0, 'cost_multiplier': 0.0}
    }
    eff = effectiveness.get(response_strategy, effectiveness['Accept'])
    response_cost = risk_emv * eff['cost_multiplier']
    risk_reduction = risk_emv * (1 - (1 - eff['probability_reduction']) * (1 - eff['impact_reduction']))
    net_benefit = risk_reduction - response_cost
    return {**eff, 'response_cost': response_cost, 'risk_reduction': risk_reduction, 'net_benefit': net_benefit}

def calculate_mitigation_costs(risk_df, direct_cost, indirect_rate, project_duration):
    """Calculate mitigation costs and risk reduction benefits"""
    mitigation_costs = {}
    total_mitigation_cost = 0
    total_risk_reduction = 0
    
    for _, row in risk_df.iterrows():
        strategy = row.get('Response Strategy', 'Accept')
        _, risk_impact = calculate_correct_emv(row, direct_cost, indirect_rate, project_duration)
        cost_factors = {'Avoid': 0.2, 'Mitigate': 0.1, 'Transfer': 0.08, 'Accept': 0.0}
        cost = risk_impact * cost_factors.get(strategy, 0)
        risk_reduction = risk_impact * (0.7 if strategy in ['Avoid', 'Mitigate', 'Transfer'] else 0)
        mitigation_costs[row['Risk Name']] = {'cost': cost, 'risk_reduction': risk_reduction, 'net_benefit': risk_reduction - cost}
        total_mitigation_cost += cost
        total_risk_reduction += risk_reduction
    
    return mitigation_costs, total_mitigation_cost, total_risk_reduction

# ============================================================================
# PROGRESS CALLBACK WITH THROTTLING
# ============================================================================
class ThrottledProgressCallback:
    def __init__(self, progress_bar, status_text, iterations, update_interval_pct=0.05):
        self.progress_bar = progress_bar
        self.status_text = status_text
        self.iterations = iterations
        self.update_interval_pct = update_interval_pct
        self.last_update = -1
        self.start_time = time.time()
        self.last_percent = -1
    
    def __call__(self, progress):
        current_percent = int(progress * 100)
        current_update_threshold = int(progress / self.update_interval_pct)
        
        if current_update_threshold > self.last_update or progress >= 1.0:
            self.last_update = current_update_threshold
            elapsed = time.time() - self.start_time
            if progress > 0:
                eta = (elapsed / progress) * (1 - progress)
                if current_percent - self.last_percent >= 2 or progress >= 1.0:
                    self.last_percent = current_percent
                    self.progress_bar.progress(progress)
                    if progress < 1.0:
                        if eta < 60:
                            self.status_text.markdown(f"### Processing {self.iterations:,} iterations... {progress:.0%} • ETA: {eta:.0f}s")
                        else:
                            self.status_text.markdown(f"### Processing {self.iterations:,} iterations... {progress:.0%} • ETA: {eta/60:.1f}m")
                    else:
                        self.status_text.markdown(f"### Complete! {self.iterations:,} iterations in {elapsed:.1f}s")

# ============================================================================
# SESSION STATE INITIALIZATION
# ============================================================================
if 'risk_factors' not in st.session_state:
    st.session_state.risk_factors = pd.DataFrame({
        'Risk Name': [
            'Weather Delays', 'Labor Shortage', 'Material Supply Delay', 
            'Ground Conditions', 'Equipment Breakdown', 'War / Geopolitical Instability',
            'Pandemic Disruption', 'Supply Chain Volatility', 'Regulatory Changes',
            'Financial Volatility', 'Design Changes', 'Safety Incident', 'Extreme Event'
        ],
        'Type': ['Threat'] * 13,
        'Probability (0-1)': [
            0.08, 0.06, 0.05, 0.05, 0.04, 0.02, 0.02, 0.03, 0.02, 0.03, 0.06, 0.01, 0.01
        ],
        'Impact (0-1)': [
            0.08, 0.06, 0.08, 0.10, 0.06, 0.15, 0.12, 0.12, 0.08, 0.08, 0.10, 0.15, 0.20
        ],
        'Distribution': ['Triangular'] * 13,
        'Optimistic (0-1)': [
            0.04, 0.02, 0.04, 0.05, 0.02, 0.05, 0.04, 0.04, 0.03, 0.03, 0.04, 0.05, 0.08
        ],
        'Most Likely (0-1)': [
            0.08, 0.06, 0.08, 0.10, 0.06, 0.15, 0.12, 0.12, 0.08, 0.08, 0.10, 0.15, 0.20
        ],
        'Pessimistic (0-1)': [
            0.16, 0.12, 0.16, 0.20, 0.12, 0.30, 0.25, 0.25, 0.16, 0.16, 0.20, 0.30, 0.40
        ],
        'Correlated With': [
            '', 'Supply Chain Volatility', 'War / Geopolitical Instability',
            '', '', 'Supply Chain Volatility',
            'War / Geopolitical Instability', 'War / Geopolitical Instability', '',
            '', 'Supply Chain Volatility', '', 'Weather Delays'
        ],
        'Correlation Strength': [
            0.0, 0.7, 0.8, 0.0, 0.0, 0.8, 0.7, 0.7, 0.0, 0.0, 0.6, 0.0, 0.5
        ],
        'Response Strategy': [
            'Accept', 'Mitigate', 'Mitigate', 'Mitigate', 'Mitigate', 
            'Transfer', 'Transfer', 'Mitigate', 'Mitigate', 'Transfer', 
            'Mitigate', 'Reduce', 'Transfer'
        ],
        'Response Plan': [
            'Weather monitoring, schedule flexibility',
            'Recruitment plan, training programs',
            'Dual sourcing, safety stock',
            'Ground investigation, contingency plans',
            'Preventive maintenance, backup equipment',
            'Political risk insurance, diversify suppliers',
            'Health protocols, remote work capabilities',
            'Supplier diversification, inventory buffer',
            'Compliance monitoring, legal review',
            'Hedging, contingency funds',
            'Change management process, design freeze',
            'Safety training, PPE, inspections',
            'Emergency response plan, insurance'
        ]
    })
    st.session_state.current_mitigation_level = "Post-Mitigation"

if 'risk_thresholds' not in st.session_state:
    st.session_state.risk_thresholds = {
        'high_prob': 0.7, 'medium_prob': 0.5, 'low_prob': 0.3,
        'high_impact': 0.7, 'medium_impact': 0.5, 'low_impact': 0.3
    }

if 'simulation_results' not in st.session_state:
    st.session_state.simulation_results = None

if 'simulation_details' not in st.session_state:
    st.session_state.simulation_details = None

if 'simulation_run_flag' not in st.session_state:
    st.session_state.simulation_run_flag = False

if 'weather_data' not in st.session_state:
    st.session_state.weather_data = None

if 'location_coords' not in st.session_state:
    st.session_state.location_coords = None

if 'ai_model' not in st.session_state:
    st.session_state.ai_model = None

if 'ai_trained' not in st.session_state:
    st.session_state.ai_trained = False

if 'commodity_data' not in st.session_state:
    st.session_state.commodity_data = None

if 'p6_activities' not in st.session_state:
    st.session_state.p6_activities = None

if 'p6_wbs_summary' not in st.session_state:
    st.session_state.p6_wbs_summary = None

if 'scenario_results' not in st.session_state:
    st.session_state.scenario_results = None

if 'operation_in_progress' not in st.session_state:
    st.session_state.operation_in_progress = False

if 'file_uploader_key_counter' not in st.session_state:
    st.session_state.file_uploader_key_counter = 0

if 'auto_project_name' not in st.session_state:
    st.session_state.auto_project_name = None

if 'auto_total_cost' not in st.session_state:
    st.session_state.auto_total_cost = None

if 'auto_indirect_cost' not in st.session_state:
    st.session_state.auto_indirect_cost = None

if 'auto_project_location' not in st.session_state:
    st.session_state.auto_project_location = None

if 'auto_duration' not in st.session_state:
    st.session_state.auto_duration = None

if 'auto_direct_cost' not in st.session_state:
    st.session_state.auto_direct_cost = None

if 'auto_indirect_rate' not in st.session_state:
    st.session_state.auto_indirect_rate = None

if 'pra_data' not in st.session_state:
    st.session_state.pra_data = None

# ============================================================================
# CURRENCY AND LOCATION DATA
# ============================================================================
currency_rates = {
    "USD": 1.0, "EUR": 0.92, "GBP": 0.79, "AED": 3.67, "SAR": 3.75, 
    "EGP": 47.5, "INR": 83.5, "PKR": 278, "QAR": 3.64, "KWD": 0.31, "BHD": 0.38
}

country_city_map = {
    "United Arab Emirates": "Dubai", "Saudi Arabia": "Riyadh", "Qatar": "Doha", 
    "Kuwait": "Kuwait City", "Oman": "Muscat", "Bahrain": "Manama",
    "Egypt": "Cairo", "Jordan": "Amman", "Lebanon": "Beirut", "Iraq": "Baghdad",
    "Turkey": "Istanbul", "India": "Mumbai", "Pakistan": "Islamabad",
    "United Kingdom": "London", "France": "Paris", "Germany": "Berlin",
    "United States": "New York", "Canada": "Toronto", "Australia": "Sydney"
}

country_currency_map = {
    "United Arab Emirates": "AED", "Saudi Arabia": "SAR", "Qatar": "QAR",
    "Egypt": "EGP", "United Kingdom": "GBP", "United States": "USD",
    "India": "INR", "Pakistan": "PKR", "Turkey": "TRY"
}

currency_symbols = {
    "USD": "$", "EUR": "€", "GBP": "£", "AED": "AED", "SAR": "SAR", 
    "EGP": "EGP", "INR": "₹", "PKR": "Rs", "TRY": "TL", "QAR": "QAR",
    "KWD": "KWD", "BHD": "BHD", "OMR": "OMR", "JOD": "JOD", "CAD": "C$",
    "AUD": "A$", "CHF": "CHF", "SEK": "SEK", "NOK": "NOK", "DKK": "DKK",
    "PLN": "zł", "CZK": "Kč", "HUF": "Ft", "ILS": "₪", "MXN": "$",
    "BRL": "R$", "ARS": "$", "CLP": "$", "PEN": "S/", "COP": "$",
    "ZAR": "R", "NGN": "₦", "KES": "KSh", "GHS": "₵"
}

country_region_map = {
    "United Arab Emirates": "GCC", "Saudi Arabia": "GCC", "Qatar": "GCC",
    "Egypt": "North Africa", "Turkey": "Europe", "India": "South Asia"
}

country_options = list(country_city_map.keys())

# ============================================================================
# WEATHER AND COMMODITY API FUNCTIONS
# ============================================================================
def geocode_location_safe(location_name):
    geocode_limiter.wait_if_needed()
    try:
        cache_key = hashlib.md5(location_name.encode()).hexdigest()
        if 'geocode_cache' not in st.session_state:
            st.session_state.geocode_cache = {}
        if cache_key in st.session_state.geocode_cache:
            return st.session_state.geocode_cache[cache_key]
        url = f"https://nominatim.openstreetmap.org/search?q={location_name}&format=json&limit=1"
        headers = {'User-Agent': 'RiskAnalysisPlatform/1.0'}
        response = requests.get(url, headers=headers, timeout=10)
        if response.status_code == 200:
            data = response.json()
            if data:
                result = (float(data[0]['lat']), float(data[0]['lon']))
                st.session_state.geocode_cache[cache_key] = result
                return result
    except Exception as e:
        st.warning(f"Geocoding failed: {str(e)[:100]}")
    return None

def get_weather_safe(location_name):
    coords = geocode_location_safe(location_name)
    if coords is None:
        return None
    try:
        url = f"https://api.open-meteo.com/v1/forecast?latitude={coords[0]}&longitude={coords[1]}&current_weather=true"
        response = requests.get(url, timeout=10)
        if response.status_code == 200:
            return response.json()
    except Exception as e:
        st.warning(f"Weather API error: {str(e)[:100]}")
    return None

def simulate_weather_data(location_name):
    seed = hash(location_name) % 2**32
    random.seed(seed)
    location_lower = location_name.lower()
    if 'dubai' in location_lower or 'doha' in location_lower or 'riyadh' in location_lower:
        temp_range = (25, 45)
    elif 'london' in location_lower or 'paris' in location_lower:
        temp_range = (5, 25)
    else:
        temp_range = (15, 35)
    return {
        'current_weather': {
            'temperature': random.uniform(temp_range[0], temp_range[1]),
            'windspeed': random.uniform(0, 30),
            'winddirection': random.uniform(0, 360)
        },
        'is_simulated': True
    }

def get_free_weather(city, country):
    location_name = f"{city}, {country}"
    weather_data = get_weather_safe(location_name)
    if weather_data and 'current_weather' in weather_data:
        weather_data['is_simulated'] = False
        return weather_data
    return simulate_weather_data(location_name)

# ============================================================================
# COMMODITY PRICES FUNCTION - DYNAMIC BY COUNTRY
# ============================================================================
def get_commodity_prices(country):
    """Return commodity prices based on selected country (converted to local currency)"""
    
    # Base prices in USD (realistic market prices as of June 2026)
    base_prices = {'steel': 650, 'cement': 125, 'fuel': 90}
    
    # Country-specific market multipliers (local factors beyond exchange rate)
    country_multipliers = {
        "United Arab Emirates": {"steel": 1.05, "cement": 1.10, "fuel": 0.85},
        "Saudi Arabia": {"steel": 1.08, "cement": 1.12, "fuel": 0.80},
        "Qatar": {"steel": 1.10, "cement": 1.15, "fuel": 0.88},
        "Kuwait": {"steel": 1.07, "cement": 1.12, "fuel": 0.82},
        "Oman": {"steel": 1.12, "cement": 1.08, "fuel": 0.90},
        "Bahrain": {"steel": 1.09, "cement": 1.10, "fuel": 0.87},
        "Egypt": {"steel": 1.20, "cement": 1.15, "fuel": 1.10},
        "Turkey": {"steel": 1.15, "cement": 1.10, "fuel": 1.15},
        "India": {"steel": 1.10, "cement": 1.08, "fuel": 1.08},
        "Pakistan": {"steel": 1.15, "cement": 1.12, "fuel": 1.12},
        "United Kingdom": {"steel": 1.10, "cement": 1.08, "fuel": 1.15},
        "United States": {"steel": 1.00, "cement": 1.00, "fuel": 1.00},
        "Germany": {"steel": 1.08, "cement": 1.05, "fuel": 1.20},
        "France": {"steel": 1.08, "cement": 1.05, "fuel": 1.18},
        "Australia": {"steel": 1.12, "cement": 1.15, "fuel": 1.05},
        "Canada": {"steel": 1.05, "cement": 1.08, "fuel": 0.95},
        "Brazil": {"steel": 1.15, "cement": 1.12, "fuel": 1.10},
        "South Africa": {"steel": 1.10, "cement": 1.12, "fuel": 1.12},
    }
    
    # Get multiplier for selected country, default to 1.0 if not found
    multiplier = country_multipliers.get(country, {"steel": 1.0, "cement": 1.0, "fuel": 1.0})
    
    # Get currency code for the country
    currency_code = country_currency_map.get(country, "USD")
    
    # Get exchange rate from your existing currency_rates dictionary
    exchange_rate = currency_rates.get(currency_code, 1.0)
    
    # Get currency symbol
    currency_symbol = currency_symbols.get(currency_code, "$")
    
    # Calculate prices: USD base × market multiplier × exchange rate to local currency
    steel_local = base_prices['steel'] * multiplier['steel'] * exchange_rate
    cement_local = base_prices['cement'] * multiplier['cement'] * exchange_rate
    fuel_local = base_prices['fuel'] * multiplier['fuel'] * exchange_rate
    
    local_prices = {
        'steel': steel_local,
        'cement': cement_local,
        'fuel': fuel_local,
        'currency': currency_symbol,
        'currency_code': currency_code
    }
    
    return local_prices
# ============================================================================
# STATISTICAL FUNCTIONS
# ============================================================================
def calculate_confidence_intervals(data, confidence_levels=[0.5, 0.8, 0.9, 0.95]):
    results = {}
    for cl in confidence_levels:
        lower = np.percentile(data, (1 - cl) * 100 / 2)
        upper = np.percentile(data, 100 - (1 - cl) * 100 / 2)
        results[f'{cl*100:.0f}% CI'] = {'lower': lower, 'upper': upper, 'range': upper - lower}
    results['mean'] = np.mean(data)
    results['std_error'] = np.std(data) / np.sqrt(len(data))
    results['coefficient_of_variation'] = np.std(data) / np.mean(data) if np.mean(data) > 0 else 0
    return results

def latin_hypercube_sampling(n_samples, n_dimensions):
    sampler = qmc.LatinHypercube(d=n_dimensions)
    sample = sampler.random(n=n_samples)
    return sample

# ============================================================================
# RISK ANALYSIS UTILITIES
# ============================================================================
def calculate_risk_priority_score(risk_df):
    risk_df = risk_df.copy()
    risk_df['Risk Score'] = risk_df['Probability (0-1)'] * risk_df['Impact (0-1)']
    urgency_map = {
        'Geopolitical Instability': 0.9, 'Pandemic Disruption': 0.7,
        'Supply Chain Disruption': 0.8, 'Labor Shortage': 0.6, 'Weather Events': 0.5
    }
    detectability_map = {
        'Geopolitical Instability': 0.3, 'Pandemic Disruption': 0.2,
        'Supply Chain Disruption': 0.5, 'Labor Shortage': 0.7, 'Weather Events': 0.8
    }
    risk_df['Urgency'] = risk_df['Risk Name'].map(urgency_map).fillna(0.5)
    risk_df['Detectability'] = risk_df['Risk Name'].map(detectability_map).fillna(0.5)
    risk_df['Priority Score'] = (risk_df['Probability (0-1)'] * risk_df['Impact (0-1)'] * 
                                  risk_df['Urgency'] * risk_df['Detectability'])
    return risk_df.sort_values('Priority Score', ascending=False)

def calculate_cascading_effects(risk_df, triggered_risk_name, all_triggered_risks=None):
    if all_triggered_risks is None:
        all_triggered_risks = set()
    all_triggered_risks.add(triggered_risk_name)
    cascading = {}
    for _, row in risk_df.iterrows():
        correlated_with = row.get('Correlated With', '')
        if correlated_with == triggered_risk_name:
            cascading[row['Risk Name']] = min(1.0, row['Probability (0-1)'] * 1.5)
            if row['Risk Name'] not in all_triggered_risks:
                deeper = calculate_cascading_effects(risk_df, row['Risk Name'], all_triggered_risks)
                cascading.update(deeper)
    return cascading

def get_phase_dependent_probability(risk_name, project_progress_percent):
    phase_profiles = {
        'Geopolitical Instability': {0: 0.4, 0.3: 0.35, 0.6: 0.25, 0.9: 0.2},
        'Supply Chain Disruption': {0: 0.35, 0.3: 0.3, 0.6: 0.15, 0.9: 0.05},
        'Labor Shortage': {0: 0.2, 0.3: 0.3, 0.6: 0.35, 0.9: 0.15},
        'Weather Events': {0: 0.15, 0.3: 0.25, 0.6: 0.2, 0.9: 0.1}
    }
    profile = phase_profiles.get(risk_name, {0: 0.2, 1.0: 0.2})
    phases = sorted(profile.keys())
    for i in range(len(phases) - 1):
        if phases[i] <= project_progress_percent <= phases[i+1]:
            ratio = (project_progress_percent - phases[i]) / (phases[i+1] - phases[i])
            return profile[phases[i]] + ratio * (profile[phases[i+1]] - profile[phases[i]])
    return profile.get(phases[-1], 0.2)

def normalize_impact(row, direct_cost, project_duration):
    impact = row.get('Impact (0-1)', 0.0)
    return min(1.0, max(0.0, impact))

# ============================================================================
# ACTIVITY-LEVEL UTILITIES
# ============================================================================
def generate_activity_durations(activities_df, iterations):
    """Generate random activity durations using Triangular distribution."""
    n_activities = len(activities_df)
    durations_matrix = np.zeros((iterations, n_activities))
    
    for idx, row in activities_df.iterrows():
        base_duration = row['Original Duration']
        
        # Triangular distribution parameters (symmetric ±15%)
        a = max(1, base_duration * 0.85)      # Minimum (Optimistic)
        b = base_duration                      # Most Likely
        c = base_duration * 1.15               # Maximum (Pessimistic)
        
        for iter_num in range(iterations):
            U = np.random.random()
            F = (b - a) / (c - a) if (c - a) > 0 else 1
            
            if U < F:
                triangular_value = a + np.sqrt(U * (c - a) * (b - a))
            else:
                triangular_value = c - np.sqrt((1 - U) * (c - a) * (c - b))
            
            durations_matrix[iter_num, idx] = triangular_value
    
    return durations_matrix

def get_enhanced_risk_task_mapping(activities_df):
    if activities_df is None or activities_df.empty:
        return {}
    mapping = {
        "Geopolitical Instability": {"activities": {}, "desc": "Material-intensive activities"},
        "Pandemic Disruption": {"activities": {}, "desc": "All activities - workforce"},
        "Supply Chain Disruption": {"activities": {}, "desc": "Material-dependent activities"},
        "Labor Shortage": {"activities": {}, "desc": "Labor-intensive activities"},
        "Weather Events": {"activities": {}, "desc": "Outdoor activities"}
    }
    for _, activity in activities_df.iterrows():
        act_id = activity['Activity ID']
        if act_id.startswith(('A200', 'A400')):
            mapping["Geopolitical Instability"]["activities"][act_id] = 0.8
        elif act_id.startswith(('A500', 'A600')):
            mapping["Geopolitical Instability"]["activities"][act_id] = 0.6
        elif act_id.startswith(('A700', 'A800', 'A900')):
            mapping["Geopolitical Instability"]["activities"][act_id] = 0.4
        else:
            mapping["Geopolitical Instability"]["activities"][act_id] = 0.3
        if act_id.startswith(('A300', 'A600', 'A900')):
            mapping["Pandemic Disruption"]["activities"][act_id] = 0.8
        else:
            mapping["Pandemic Disruption"]["activities"][act_id] = 0.5
        if act_id.startswith(('A200', 'A400')):
            mapping["Supply Chain Disruption"]["activities"][act_id] = 0.9
        elif act_id.startswith(('A500', 'A600')):
            mapping["Supply Chain Disruption"]["activities"][act_id] = 0.7
        elif act_id.startswith(('A700', 'A800', 'A900')):
            mapping["Supply Chain Disruption"]["activities"][act_id] = 0.5
        else:
            mapping["Supply Chain Disruption"]["activities"][act_id] = 0.3
        if act_id.startswith(('A600', 'A900')):
            mapping["Labor Shortage"]["activities"][act_id] = 0.9
        elif act_id.startswith(('A500', 'A700')):
            mapping["Labor Shortage"]["activities"][act_id] = 0.5
        elif act_id.startswith(('A800',)):
            mapping["Labor Shortage"]["activities"][act_id] = 0.3
        else:
            mapping["Labor Shortage"]["activities"][act_id] = 0.2
        if act_id.startswith(('A300',)):
            mapping["Weather Events"]["activities"][act_id] = 1.0
        elif act_id.startswith(('A500',)):
            mapping["Weather Events"]["activities"][act_id] = 0.8
        elif act_id.startswith(('A700',)):
            mapping["Weather Events"]["activities"][act_id] = 0.6
        elif act_id.startswith(('A800',)):
            mapping["Weather Events"]["activities"][act_id] = 0.4
        else:
            mapping["Weather Events"]["activities"][act_id] = 0.2
    return mapping

# ============================================================================
# MONTE CARLO SIMULATION ENGINE
# ============================================================================
def run_monte_carlo_full_optimized(risk_df, base_duration, base_total_cost, indirect_rate, iterations, progress_callback=None, activities_df=None, pra_data=None):
    validation_errors = validate_risk_input(risk_df)
    if validation_errors:
        st.error(f"Input validation failed: {', '.join(validation_errors)}")
        return None, None, None, None
    
    n_risks = len(risk_df)
    lhs_samples = latin_hypercube_sampling(iterations, n_risks)
    
    # Activity uncertainty configuration
    uncertainty_pct = st.session_state.get('uncertainty_value', 15) / 100
    
    # Triangular distribution for activity uncertainty (symmetric ± uncertainty_pct%)
    U_act = np.random.random(iterations)
    a_act = base_duration * (1 - uncertainty_pct)   # Minimum (Optimistic)
    b_act = base_duration                            # Most Likely
    c_act = base_duration * (1 + uncertainty_pct)    # Maximum (Pessimistic)
    
    F_act = (b_act - a_act) / (c_act - a_act) if (c_act - a_act) > 0 else 1
    
    activity_uncertainty = np.where(
        U_act < F_act,
        a_act + np.sqrt(U_act * (c_act - a_act) * (b_act - a_act)),
        c_act - np.sqrt((1 - U_act) * (c_act - a_act) * (c_act - b_act))
    )
    base_duration_with_uncertainty = activity_uncertainty
    
    risk_names = risk_df['Risk Name'].tolist()
    risk_types = risk_df['Type'].tolist()
    base_probs = risk_df['Probability (0-1)'].values
    optimistics = risk_df['Optimistic (0-1)'].values
    most_likelys = risk_df['Most Likely (0-1)'].values
    pessimistics = risk_df['Pessimistic (0-1)'].values
    distributions = risk_df['Distribution'].tolist()
    response_strategies = risk_df.get('Response Strategy', ['Accept'] * n_risks).tolist()
    
    effectiveness_list = []
    for idx in range(n_risks):
        emv, _ = calculate_correct_emv(risk_df.iloc[idx], base_total_cost, indirect_rate, base_duration)
        effectiveness_list.append(get_response_effectiveness(response_strategies[idx], emv))
    
    durations = []
    costs = []
    risk_contributions = {name: [] for name in risk_names}
    iteration_details = []
    
    batch_size = 500
    num_batches = (iterations + batch_size - 1) // batch_size
    last_progress = 0
    
    for batch_idx in range(num_batches):
        batch_start = batch_idx * batch_size
        batch_end = min(batch_start + batch_size, iterations)
        for local_idx in range(batch_end - batch_start):
            global_idx = batch_start + local_idx
            duration_factor = 1.0
            cost_factor = 1.0
            iter_risks = {}
            project_progress = global_idx / iterations
            triggered_cascading = {}
            
            for idx in range(n_risks):
                risk_name = risk_names[idx]
                base_prob = base_probs[idx]
                opt = optimistics[idx]
                ml = most_likelys[idx]
                pess = pessimistics[idx]
                dist = distributions[idx]
                risk_type = risk_types[idx]
                effectiveness = effectiveness_list[idx]
                
                adjusted_prob = base_prob * (1 - effectiveness['probability_reduction'])
                phase_prob = get_phase_dependent_probability(risk_name, project_progress)
                if base_prob > 0:
                    adjusted_prob = adjusted_prob * max(0.5, phase_prob / base_prob)
                adjusted_prob = min(1.0, max(0.0, adjusted_prob))
                if risk_name in triggered_cascading:
                    adjusted_prob = min(1.0, adjusted_prob * triggered_cascading[risk_name])
                
                random_value = lhs_samples[global_idx, idx]
                if random_value < adjusted_prob:
                    if dist == "Triangular":
                        impact = np.random.triangular(opt, ml, pess)
                    elif dist == "Uniform":
                        impact = np.random.uniform(opt, pess)
                    else:
                        impact = (opt + 4*ml + pess) / 6
                        impact = np.random.normal(impact, (pess - opt) / 6)
                    impact = max(opt, min(pess, impact))
                    impact *= (1 - effectiveness['impact_reduction'])
                    cascading = calculate_cascading_effects(risk_df, risk_name)
                    for casc_risk, casc_factor in cascading.items():
                        if casc_risk not in triggered_cascading:
                            triggered_cascading[casc_risk] = casc_factor
                    if risk_type == 'Threat':
                        # Cap individual risk impact at 0.5 (50% increase max)
                        impact_multiplier = 1 + min(impact, 0.5)
                        duration_factor *= impact_multiplier
                        cost_factor = min(cost_factor * impact_multiplier, 1.12)
                        iter_risks[risk_name] = (impact_multiplier - 1) * 100
                    else:
                        impact_multiplier = 1 - min(impact, 0.5)
                        duration_factor *= max(0.5, impact_multiplier)
                        cost_factor *= max(0.5, impact_multiplier)
                        iter_risks[risk_name] = (1 - impact_multiplier) * 100
                else:
                    iter_risks[risk_name] = 0
                risk_contributions[risk_name].append(iter_risks.get(risk_name, 0))
            
            # Cap total duration factor to prevent unrealistic compounding
            MAX_DURATION_FACTOR = 1.10
            if duration_factor > MAX_DURATION_FACTOR:
                scale_ratio = MAX_DURATION_FACTOR / duration_factor
                duration_factor = MAX_DURATION_FACTOR
                cost_factor = 1 + (cost_factor - 1) * scale_ratio
            
            total_duration = base_duration_with_uncertainty[global_idx] * duration_factor
            total_cost = base_total_cost * cost_factor
            
            durations.append(total_duration)
            costs.append(total_cost)
            
            if len(iteration_details) < 5000:
                iteration_details.append({
                    'Iteration': global_idx + 1, 
                    'Duration (days)': round(total_duration, 1), 
                    'Cost': round(total_cost, 0)
                })
            
            if progress_callback and global_idx - last_progress >= max(1, iterations // 100):
                progress_callback(global_idx / iterations)
                last_progress = global_idx
    
    if progress_callback:
        progress_callback(1.0)
    return np.array(durations), np.array(costs), risk_contributions, pd.DataFrame(iteration_details)

# ============================================================================
# AI RISK PREDICTION MODEL
# ============================================================================
@st.cache_resource
def train_risk_prediction_model(risk_df):
    np.random.seed(42)
    n_samples = 10000
    n_risks = len(risk_df)
    X = np.random.rand(n_samples, n_risks)
    y = np.zeros(n_samples)
    for i in range(n_samples):
        delay = 0
        for j, (_, row) in enumerate(risk_df.iterrows()):
            prob = row['Probability (0-1)']
            delay += X[i, j] * prob * 0.5
            correlated_with = row.get('Correlated With', '')
            if correlated_with:
                for k, (_, row2) in enumerate(risk_df.iterrows()):
                    if row2['Risk Name'] == correlated_with:
                        delay += X[i, j] * X[i, k] * 0.3
                        break
        y[i] = min(0.95, delay)
    model = RandomForestRegressor(n_estimators=100, random_state=42, n_jobs=-1)
    scaler = StandardScaler()
    X_scaled = scaler.fit_transform(X)
    model.fit(X_scaled, y)
    return model, scaler

# ============================================================================
# EXPORT FUNCTIONS
# ============================================================================
def create_professional_excel(df, sheet_name, title):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=2)
        workbook = writer.book
        worksheet = writer.sheets[sheet_name]
        worksheet.merge_cells('A1:Z1')
        title_cell = worksheet['A1']
        title_cell.value = title
        title_cell.font = Font(name='Calibri', size=14, bold=True, color='000000')
        title_cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        worksheet.merge_cells('A2:Z2')
        date_cell = worksheet['A2']
        date_cell.value = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        date_cell.font = Font(name='Calibri', size=10, italic=True, color='666666')
        date_cell.alignment = Alignment(horizontal='center', vertical='center')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        for col in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=3, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
        data_font = Font(name='Calibri', size=10, color='000000')
        thin_border = Border(left=Side(style='thin', color='CCCCCC'), right=Side(style='thin', color='CCCCCC'), top=Side(style='thin', color='CCCCCC'), bottom=Side(style='thin', color='CCCCCC'))
        for row in range(4, len(df) + 4):
            for col in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row, column=col)
                cell.font = data_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border
                if row % 2 == 0:
                    cell.fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
            cell = worksheet.cell(row=row, column=1)
            cell.font = Font(name='Calibri', size=10, bold=True, color='000000')
        for col in range(1, len(df.columns) + 1):
            try:
                max_length = max(df.iloc[:, col-1].astype(str).apply(len).max(), len(str(df.columns[col-1]))) + 2
                worksheet.column_dimensions[get_column_letter(col)].width = min(max_length, 50)
            except:
                worksheet.column_dimensions[get_column_letter(col)].width = 15
    output.seek(0)
    return output

def create_zip_with_charts(durations, costs, risk_contributions, user_percentiles, percentiles, cost_percentiles):
    import matplotlib.pyplot as plt
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
        # Duration histogram
        plt.figure(figsize=(12, 6))
        plt.hist(durations, bins=50, color='#3b82f6', alpha=0.7, edgecolor='black', label='Iterations')
        colors = ['#ef4444', '#f97316', '#eab308', '#10b981', '#06b6d4', '#8b5cf6', '#ec4899', '#f43f5e', '#14b8a6']
        for idx, (name, value) in enumerate(percentiles.items()):
            plt.axvline(x=value, color=colors[idx % len(colors)], linestyle='--', linewidth=2, label=f'{name}: {value:.0f} days')
        plt.xlabel('Days')
        plt.ylabel('Frequency')
        plt.title('Project Duration Distribution')
        plt.legend(loc='upper right', framealpha=0.9)
        plt.tight_layout()
        plt.savefig('temp_duration.png', dpi=150, facecolor='white')
        plt.close()
        with open('temp_duration.png', 'rb') as f:
            zip_file.writestr('duration_distribution.png', f.read())
        
        # S-Curve
        sorted_indices = np.argsort(durations)
        sorted_durations = durations[sorted_indices]
        cumulative = np.arange(1, len(sorted_durations) + 1) / (len(sorted_durations) + 1)
        extended_durations = np.insert(sorted_durations, 0, sorted_durations[0] * 0.95)
        extended_cumulative = np.insert(cumulative, 0, 0)
        plt.figure(figsize=(12, 6))
        plt.plot(extended_durations, extended_cumulative, color='#3b82f6', linewidth=3, label='Cumulative Probability')
        for idx, (name, value) in enumerate(percentiles.items()):
            prob = int(name.replace("P", "")) / 100
            plt.axhline(y=prob, color=colors[idx % len(colors)], linestyle='--', linewidth=2, label=f'{name}: {value:.0f} days')
        plt.xlabel('Days')
        plt.ylabel('Probability (0-1)')
        plt.title('Cumulative Probability Curve (S-Curve)')
        plt.legend(loc='lower right', framealpha=0.9)
        plt.tight_layout()
        plt.savefig('temp_scurve.png', dpi=150, facecolor='white')
        plt.close()
        with open('temp_scurve.png', 'rb') as f:
            zip_file.writestr('s_curve.png', f.read())
        
        # Cost histogram
        plt.figure(figsize=(12, 6))
        plt.hist(costs, bins=50, color='#10b981', alpha=0.7, edgecolor='black', label='Iterations')
        for idx, (name, value) in enumerate(cost_percentiles.items()):
            plt.axvline(x=value, color=colors[idx % len(colors)], linestyle='--', linewidth=2, label=f'{name}: ${value:,.0f}')
        plt.xlabel('Cost')
        plt.ylabel('Frequency')
        plt.title('Project Cost Distribution')
        plt.legend(loc='upper right', framealpha=0.9)
        plt.tight_layout()
        plt.savefig('temp_cost.png', dpi=150, facecolor='white')
        plt.close()
        with open('temp_cost.png', 'rb') as f:
            zip_file.writestr('cost_distribution.png', f.read())
        
        # Tornado chart
        sensitivity = []
        for risk_name, contributions in risk_contributions.items():
            if len(contributions) > 0:
                sensitivity.append({'Risk': risk_name, 'Impact (%)': np.mean(np.abs(contributions))})
        sensitivity_df = pd.DataFrame(sensitivity).sort_values('Impact (%)', ascending=True)
        plt.figure(figsize=(10, max(6, len(sensitivity_df) * 0.4)))
        bars = plt.barh(sensitivity_df['Risk'], sensitivity_df['Impact (%)'], color='#3b82f6', label='Impact on Duration')
        plt.xlabel('Impact on Duration (%)')
        plt.ylabel('Risk Factor')
        plt.title('Risk Impact Ranking (Tornado Chart)')
        plt.legend([bars], ['Impact on Duration (%)'], loc='lower right')
        plt.tight_layout()
        plt.savefig('temp_tornado.png', dpi=150, facecolor='white')
        plt.close()
        with open('temp_tornado.png', 'rb') as f:
            zip_file.writestr('tornado_chart.png', f.read())
        
        # Pie chart
        risk_df_pie = pd.DataFrame(sensitivity).sort_values('Impact (%)', ascending=False).head(10)
        plt.figure(figsize=(10, 8))
        plt.pie(risk_df_pie['Impact (%)'], labels=risk_df_pie['Risk'], autopct='%1.1f%%', startangle=90)
        plt.title('Top 10 Risk Contributors')
        plt.legend(risk_df_pie['Risk'], loc='upper right', bbox_to_anchor=(1.3, 1))
        plt.tight_layout()
        plt.savefig('temp_pie.png', dpi=150, facecolor='white')
        plt.close()
        with open('temp_pie.png', 'rb') as f:
            zip_file.writestr('risk_contributions.png', f.read())
        
        readme_content = f"""Risk Analysis Platform - Chart Export
Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
Iterations: {len(durations)}
Percentiles: {', '.join([str(p) for p in user_percentiles])}

Files included:
- duration_distribution.png: Histogram of project duration (with percentile lines)
- s_curve.png: Cumulative probability curve (with percentile lines)
- cost_distribution.png: Histogram of project cost (with percentile lines)
- tornado_chart.png: Sensitivity analysis
- risk_contributions.png: Top 10 risk contributors pie chart
"""
        zip_file.writestr('readme.txt', readme_content)
    zip_buffer.seek(0)
    return zip_buffer

# ============================================================================
# CONVERGENCE AND SCENARIO FUNCTIONS
# ============================================================================
def check_convergence(durations, window=1000):
    if len(durations) < window * 2:
        return None, None
    p50_convergence = []
    for i in range(window, len(durations), window):
        p50_convergence.append(np.percentile(durations[:i], 50))
    if len(p50_convergence) >= 3:
        recent = p50_convergence[-3:]
        variation = np.std(recent) / np.mean(recent) if np.mean(recent) > 0 else 1.0
        return variation, p50_convergence
    return None, None

def run_scenario_comparison(factor, risk_df, base_duration, base_total_cost, indirect_rate, iterations=1000, pra_data=None):
    durations, costs, _, _ = run_monte_carlo_full_optimized(
        risk_df, base_duration * factor, base_total_cost * factor, indirect_rate, iterations, None, None, pra_data
    )
    if durations is None:
        return None
    return {
        'p50': np.percentile(durations, 50), 
        'p80': np.percentile(durations, 80), 
        'mean': np.mean(durations), 
        'std': np.std(durations), 
        'cost_p50': np.percentile(costs, 50), 
        'cost_p80': np.percentile(costs, 80)
    }

# ============================================================================
# RISK TASK MAPPING
# ============================================================================
def get_risk_task_mapping(activities_df=None):
    if activities_df is not None and not activities_df.empty:
        all_ids = activities_df['Activity ID'].tolist()
    else:
        all_ids = []
    mapping = {
        "Geopolitical Instability": {
            "prefixes": ["A200", "A400", "A500", "A600", "A700", "A800", "A900"], 
            "desc": "Material-intensive activities", 
            "affected_count": 0
        },
        "Pandemic Disruption": {
            "prefixes": ["ALL"], 
            "desc": "All activities", 
            "affected_count": 0
        },
        "Supply Chain Disruption": {
            "prefixes": ["A200", "A400", "A500", "A600", "A700", "A800", "A900"], 
            "desc": "Material-dependent activities", 
            "affected_count": 0
        },
        "Labor Shortage": {
            "prefixes": ["A600", "A900"], 
            "desc": "Labor-intensive activities", 
            "affected_count": 0
        },
        "Weather Events": {
            "prefixes": ["A300", "A500", "A700", "A800"], 
            "desc": "Outdoor activities", 
            "affected_count": 0
        }
    }
    if all_ids:
        for risk_name, risk_info in mapping.items():
            if risk_info.get("prefixes") == ["ALL"]:
                risk_info["affected_ids"] = all_ids
                risk_info["affected_count"] = len(all_ids)
            else:
                affected = []
                for prefix in risk_info["prefixes"]:
                    matched = [aid for aid in all_ids if aid.startswith(prefix)]
                    affected.extend(matched)
                risk_info["affected_ids"] = list(set(affected))
                risk_info["affected_count"] = len(risk_info["affected_ids"])
    else:
        for risk_name, risk_info in mapping.items():
            risk_info["affected_ids"] = []
            risk_info["affected_count"] = 0
    return mapping

# ============================================================================
# PRIMAVERA P6 PARSER FUNCTIONS
# ============================================================================
def parse_excel_file(file_content):
    """Parse Primavera P6 exported Excel file"""
    try:
        excel_file = io.BytesIO(file_content)
        task_df = pd.read_excel(excel_file, sheet_name='TASK', header=1)
        
        task_df.columns = [
            'Activity ID', 'Status', 'WBS Code', 'Budgeted Cost', 
            'Total Float', 'Activity Name', 'Critical', 'Original Duration', 
            'Delete Row'
        ]
        
        task_df['Budgeted Cost'] = task_df['Budgeted Cost'].astype(str).str.replace('$', '').str.replace(',', '').str.strip()
        task_df['Budgeted Cost'] = pd.to_numeric(task_df['Budgeted Cost'], errors='coerce').fillna(0)
        task_df['Original Duration'] = pd.to_numeric(task_df['Original Duration'], errors='coerce').fillna(0).astype(int)
        
        # Calculate total cost
        direct_cost_mask = task_df['Activity ID'] != 'Indirect'
        direct_cost = task_df[direct_cost_mask]['Budgeted Cost'].sum()
        
        indirect_activity = task_df[task_df['Activity ID'] == 'Indirect']
        indirect_cost = 0
        if len(indirect_activity) > 0:
            indirect_cost = float(indirect_activity['Budgeted Cost'].iloc[0])
        
        total_cost = direct_cost + indirect_cost
        
        # Calculate project duration
        total_duration = 0
        
        try:
            try:
                taskrsrc_df = pd.read_excel(excel_file, sheet_name='TASKRSRC', header=1)
                taskrsrc_df.columns = ['Resource ID', 'Activity ID', 'Status', 'Role ID', 
                                       'Cost Account ID', 'Resource Type', 'Start', 'Finish', 'Delete Row']
                taskrsrc_df['Start'] = pd.to_datetime(taskrsrc_df['Start'], errors='coerce')
                taskrsrc_df['Finish'] = pd.to_datetime(taskrsrc_df['Finish'], errors='coerce')
                min_start = taskrsrc_df['Start'].min()
                max_finish = taskrsrc_df['Finish'].max()
                if pd.notna(min_start) and pd.notna(max_finish):
                    total_duration = (max_finish - min_start).days
            except:
                pass
            
            if total_duration == 0:
                critical_activities = task_df[task_df['Critical'] == 'Y']
                critical_activities = critical_activities[~critical_activities['Activity ID'].isin(['Start', 'Finish', 'Indirect'])]
                total_duration = critical_activities['Original Duration'].sum()
            
            if total_duration == 0:
                start_activity = task_df[task_df['Activity ID'] == 'Start']
                finish_activity = task_df[task_df['Activity ID'] == 'Finish']
                if len(start_activity) > 0 and len(finish_activity) > 0:
                    finish_dur = finish_activity['Original Duration'].iloc[0]
                    start_dur = start_activity['Original Duration'].iloc[0]
                    if finish_dur > start_dur:
                        total_duration = finish_dur - start_dur
            
            if total_duration == 0 and len(indirect_activity) > 0:
                total_duration = indirect_activity['Original Duration'].iloc[0]
            
            if total_duration <= 0:
                total_duration = 1693
                
        except Exception as e:
            st.warning(f"Could not read duration from file: {e}")
            total_duration = 1693
        
        # Build activities dataframe
        activities_df = task_df[~task_df['Activity ID'].isin(['Start', 'Finish', 'Indirect'])].copy()
        activities_df = activities_df[['Activity ID', 'Activity Name', 'Original Duration', 
                                        'Budgeted Cost', 'WBS Code', 'Status', 'Critical']]
        activities_df.rename(columns={'Budgeted Cost': 'target_cost'}, inplace=True)
        
        # Project name
        project_name = "Underground Metro Station"
        unique_wbs = activities_df['WBS Code'].dropna().unique()
        for wbs in unique_wbs:
            if '.' not in str(wbs) and len(str(wbs)) > 0 and str(wbs) != 'UMS.0':
                project_name = str(wbs)
                break
        
        return {
            'activities': activities_df,
            'total_duration': total_duration,
            'total_cost': total_cost,
            'indirect_cost': indirect_cost,
            'project_name': project_name,
            'project_location': 'Middle East'
        }
        
    except Exception as e:
        st.error(f"Error parsing Excel file: {str(e)}")
        return None

def extract_calendar_hours_per_day(root, namespaces):
    try:
        project_elem = root.find('.//p6:Project', namespaces)
        if project_elem is not None:
            calendar_id_elem = project_elem.find('p6:ActivityDefaultCalendarObjectId', namespaces)
            if calendar_id_elem is not None and calendar_id_elem.text:
                calendar_elem = root.find(f'.//p6:Calendar[p6:ObjectId="{calendar_id_elem.text}"]', namespaces)
                if calendar_elem is not None:
                    hours_elem = calendar_elem.find('p6:HoursPerDay', namespaces)
                    if hours_elem is not None and hours_elem.text:
                        return float(hours_elem.text)
        calendar_elem = root.find('.//p6:Calendar[p6:IsDefault="1"]', namespaces)
        if calendar_elem is not None:
            hours_elem = calendar_elem.find('p6:HoursPerDay', namespaces)
            if hours_elem is not None and hours_elem.text:
                return float(hours_elem.text)
    except:
        pass
    return 8.0

def extract_wbs_summary(root, namespaces):
    wbs_elements = {}
    for wbs in root.findall('.//p6:WBS', namespaces):
        obj_id = wbs.find('p6:ObjectId', namespaces)
        name = wbs.find('p6:Name', namespaces)
        code = wbs.find('p6:Code', namespaces)
        parent_id = wbs.find('p6:ParentObjectId', namespaces)
        if obj_id is not None and name is not None and name.text:
            wbs_elements[obj_id.text] = {
                'name': name.text, 
                'code': code.text if code is not None else '', 
                'parent_id': parent_id.text if parent_id is not None else None, 
                'children': [], 
                'total_duration': 0, 
                'total_cost': 0
            }
    for wbs_id, wbs in wbs_elements.items():
        if wbs['parent_id'] and wbs['parent_id'] in wbs_elements:
            wbs_elements[wbs['parent_id']]['children'].append(wbs_id)
    
    root_wbs_id = None
    for wbs_id, wbs in wbs_elements.items():
        if wbs['parent_id'] is None or wbs['parent_id'] == '':
            root_wbs_id = wbs_id
            break
    
    hours_per_day = extract_calendar_hours_per_day(root, namespaces)
    for activity in root.findall('.//p6:Activity', namespaces):
        wbs_obj_id = activity.find('p6:WBSObjectId', namespaces)
        if wbs_obj_id is None or wbs_obj_id.text not in wbs_elements:
            continue
        labor = activity.find('p6:AtCompletionLaborCost', namespaces)
        nonlabor = activity.find('p6:AtCompletionNonLaborCost', namespaces)
        expense = activity.find('p6:AtCompletionExpenseCost', namespaces)
        cost = 0
        if labor is not None and labor.text:
            try: cost += float(labor.text)
            except: pass
        if nonlabor is not None and nonlabor.text:
            try: cost += float(nonlabor.text)
            except: pass
        if expense is not None and expense.text:
            try: cost += float(expense.text)
            except: pass
        planned_duration = activity.find('p6:PlannedDuration', namespaces)
        duration = 0
        if planned_duration is not None and planned_duration.text:
            try: duration = float(planned_duration.text) / hours_per_day
            except: pass
        wbs_elements[wbs_obj_id.text]['total_duration'] += duration
        wbs_elements[wbs_obj_id.text]['total_cost'] += cost
    
    def rollup_totals(wbs_id):
        wbs = wbs_elements[wbs_id]
        for child_id in wbs['children']:
            if child_id in wbs_elements:
                child = rollup_totals(child_id)
                wbs['total_duration'] += child['total_duration']
                wbs['total_cost'] += child['total_cost']
        return wbs
    
    if root_wbs_id and root_wbs_id in wbs_elements:
        root_wbs = rollup_totals(root_wbs_id)
        return {
            'project_name': root_wbs['name'], 
            'total_duration': int(round(root_wbs['total_duration'])), 
            'total_cost': root_wbs['total_cost'], 
            'wbs_hierarchy': wbs_elements, 
            'root_wbs_id': root_wbs_id
        }
    return None

def extract_project_from_element(root, namespaces):
    project_elem = root.find('.//p6:Project', namespaces)
    if project_elem is None:
        return None
    name_elem = project_elem.find('p6:Name', namespaces)
    project_name = name_elem.text if name_elem is not None else "Unknown"
    start_elem = project_elem.find('p6:PlannedStartDate', namespaces)
    finish_elem = project_elem.find('p6:ScheduledFinishDate', namespaces)
    total_duration = 1360
    if start_elem is not None and finish_elem is not None and start_elem.text and finish_elem.text:
        try:
            start_str = start_elem.text.replace('Z', '').split('T')[0]
            finish_str = finish_elem.text.replace('Z', '').split('T')[0]
            start_date = datetime.fromisoformat(start_str)
            finish_date = datetime.fromisoformat(finish_str)
            total_duration = (finish_date - start_date).days
        except:
            pass
    project_location = "Middle East"
    desc_elem = project_elem.find('p6:Description', namespaces)
    if desc_elem is not None and desc_elem.text:
        desc_text = desc_elem.text.lower()
        if 'dubai' in desc_text or 'uae' in desc_text:
            project_location = "United Arab Emirates"
        elif 'riyadh' in desc_text or 'saudi' in desc_text:
            project_location = "Saudi Arabia"
        elif 'doha' in desc_text or 'qatar' in desc_text:
            project_location = "Qatar"
    return {
        'project_name': project_name, 
        'project_location': project_location, 
        'total_duration': total_duration, 
        'total_cost': None
    }

def parse_xer_file(file_content, file_type="xer"):
    """Parse Primavera P6 XER or XML file"""
    try:
        if file_type == "xml":
            content_str = file_content.decode('utf-8', errors='ignore')
            root = ET.fromstring(content_str)
            namespaces_list = [
                {'p6': 'http://xmlns.oracle.com/Primavera/P6Professional/V22.12/API/BusinessObjects'},
                {'p6': 'http://xmlns.oracle.com/Primavera/P6Professional/V20.12/API/BusinessObjects'},
                {'p6': 'http://xmlns.oracle.com/Primavera/P6Enterprise/V1/API/BusinessObjects'},
                {'': ''}
            ]
            hours_per_day = 8
            for namespaces in namespaces_list:
                if namespaces.get('p6', ''):
                    hours_per_day = extract_calendar_hours_per_day(root, namespaces)
                    if hours_per_day != 8:
                        break
            activities = []
            wbs_summary = None
            project_params = None
            
            for namespaces in namespaces_list:
                ns = namespaces
                if ns.get('p6', ''):
                    for activity_elem in root.findall('.//p6:Activity', ns):
                        id_elem = activity_elem.find('p6:Id', ns)
                        activity_id = id_elem.text if id_elem is not None and id_elem.text else ''
                        name_elem = activity_elem.find('p6:Name', ns)
                        activity_name = name_elem.text if name_elem is not None and name_elem.text else ''
                        if not activity_id or not activity_name:
                            continue
                        planned_duration = 0
                        planned_duration_elem = activity_elem.find('p6:PlannedDuration', ns)
                        if planned_duration_elem is not None and planned_duration_elem.text:
                            try: planned_duration = float(planned_duration_elem.text) / hours_per_day
                            except: pass
                        at_completion_duration = 0
                        at_completion_elem = activity_elem.find('p6:AtCompletionDuration', ns)
                        if at_completion_elem is not None and at_completion_elem.text:
                            try: at_completion_duration = float(at_completion_elem.text) / hours_per_day
                            except: pass
                        duration = max(planned_duration, at_completion_duration)
                        duration = max(1, int(round(duration))) if duration > 0 else 1
                        budgeted_cost = 0
                        labor_cost_elem = activity_elem.find('p6:AtCompletionLaborCost', ns)
                        if labor_cost_elem is not None and labor_cost_elem.text:
                            try: budgeted_cost += float(labor_cost_elem.text)
                            except: pass
                        nonlabor_cost_elem = activity_elem.find('p6:AtCompletionNonLaborCost', ns)
                        if nonlabor_cost_elem is not None and nonlabor_cost_elem.text:
                            try: budgeted_cost += float(nonlabor_cost_elem.text)
                            except: pass
                        expense_cost_elem = activity_elem.find('p6:AtCompletionExpenseCost', ns)
                        if expense_cost_elem is not None and expense_cost_elem.text:
                            try: budgeted_cost += float(expense_cost_elem.text)
                            except: pass
                        if budgeted_cost == 0:
                            planned_labor = activity_elem.find('p6:PlannedLaborCost', ns)
                            if planned_labor is not None and planned_labor.text:
                                try: budgeted_cost += float(planned_labor.text)
                                except: pass
                            planned_nonlabor = activity_elem.find('p6:PlannedNonLaborCost', ns)
                            if planned_nonlabor is not None and planned_nonlabor.text:
                                try: budgeted_cost += float(planned_nonlabor.text)
                                except: pass
                            planned_expense = activity_elem.find('p6:PlannedExpenseCost', ns)
                            if planned_expense is not None and planned_expense.text:
                                try: budgeted_cost += float(planned_expense.text)
                                except: pass
                        start_date = ''
                        start_elem = activity_elem.find('p6:StartDate', ns)
                        if start_elem is not None and start_elem.text:
                            start_date = start_elem.text[:10] if len(start_elem.text) > 10 else start_elem.text
                        finish_date = ''
                        finish_elem = activity_elem.find('p6:FinishDate', ns)
                        if finish_elem is not None and finish_elem.text:
                            finish_date = finish_elem.text[:10] if len(finish_elem.text) > 10 else finish_elem.text
                        wbs_id = ''
                        wbs_elem = activity_elem.find('p6:WBSObjectId', ns)
                        if wbs_elem is not None and wbs_elem.text:
                            wbs_id = wbs_elem.text
                        status = ''
                        status_elem = activity_elem.find('p6:Status', ns)
                        if status_elem is not None and status_elem.text:
                            status = status_elem.text
                        activity_type = ''
                        type_elem = activity_elem.find('p6:Type', ns)
                        if type_elem is not None and type_elem.text:
                            activity_type = type_elem.text
                        wbs_name = ''
                        wbs_obj_id_elem = activity_elem.find('p6:WBSObjectId', ns)
                        if wbs_obj_id_elem is not None and wbs_obj_id_elem.text:
                            for wbs_elem_search in root.findall('.//p6:WBS', ns):
                                obj_id_elem = wbs_elem_search.find('p6:ObjectId', ns)
                                if obj_id_elem is not None and obj_id_elem.text == wbs_obj_id_elem.text:
                                    name_elem_wbs = wbs_elem_search.find('p6:Name', ns)
                                    if name_elem_wbs is not None and name_elem_wbs.text:
                                        wbs_name = name_elem_wbs.text
                                    break
                        activities.append({
                            'Activity ID': activity_id, 
                            'Activity Name': activity_name, 
                            'Original Duration': duration, 
                            'Budgeted Cost': budgeted_cost if budgeted_cost > 0 else 10000,
                            'Start Date': start_date, 
                            'Finish Date': finish_date, 
                            'WBS ID': wbs_id, 
                            'WBS Name': wbs_name, 
                            'Status': status, 
                            'Type': activity_type
                        })
                if activities:
                    wbs_summary = extract_wbs_summary(root, ns)
                    project_params = extract_project_from_element(root, ns)
                    break
            
            if activities:
                df = pd.DataFrame(activities)
                df['Original Duration'] = df['Original Duration'].fillna(1).astype(int)
                df['Original Duration'] = df['Original Duration'].clip(lower=1)
                df = df.sort_values('Activity ID').reset_index(drop=True)
                result = {'activities': df}
                if wbs_summary:
                    result['wbs_summary'] = wbs_summary
                if project_params:
                    result['project_params'] = project_params
                    result['total_duration'] = project_params['total_duration']
                    result['total_cost'] = wbs_summary['total_cost'] if wbs_summary and wbs_summary.get('total_cost') else df['Budgeted Cost'].sum()
                return result
            else:
                st.error("No activities found in XML file.")
                return None
        else:
            # XER file parsing
            decoded_content = None
            for encoding in ['utf-8', 'latin-1', 'cp1252', 'iso-8859-1']:
                try:
                    decoded_content = file_content.decode(encoding, errors='ignore')
                    break
                except:
                    continue
            if not decoded_content:
                decoded_content = file_content.decode('utf-8', errors='ignore')
            
            lines = decoded_content.split('\n')
            in_activities = False
            headers = []
            activities = []
            
            for line in lines:
                line = line.strip()
                if not line:
                    continue
                if line.startswith('%T') and ('task' in line.lower() or 'activity' in line.lower()):
                    in_activities = True
                    headers = line[2:].split('\t')
                    headers = [h.strip().replace('?', '').replace('(', '').replace(')', '').replace(' ', '_') for h in headers]
                    continue
                if in_activities and line.startswith('%R'):
                    values = line[2:].split('\t')
                    if len(values) >= 2:
                        activity = {}
                        for i, header in enumerate(headers):
                            if i < len(values):
                                activity[header] = values[i]
                        task_code = activity.get('task_code', '') or activity.get('task_id', '')
                        task_name = activity.get('task_name', '') or activity.get('activity_name', '')
                        if task_code and task_name:
                            try:
                                duration_hours = activity.get('target_drtn_hr_cnt', '0') or activity.get('target_duration_hr_cnt', '0')
                                if duration_hours:
                                    duration_days = max(1, round(float(duration_hours) / 8.0))
                                else:
                                    duration_days = 1
                            except:
                                duration_days = 1
                            try:
                                budget_cost = float(activity.get('total_cost', 0)) or float(activity.get('total_cost_amt', 0))
                                if budget_cost == 0:
                                    budget_cost = 10000
                            except:
                                budget_cost = 10000
                            activities.append({
                                'Activity ID': task_code, 
                                'Activity Name': task_name, 
                                'Original Duration': duration_days, 
                                'Budgeted Cost': budget_cost,
                                'Start Date': activity.get('plan_start_date', ''), 
                                'Finish Date': activity.get('plan_end_date', ''), 
                                'WBS ID': activity.get('wbs_id', ''), 
                                'WBS Name': '', 
                                'Status': activity.get('status_code', ''), 
                                'Type': ''
                            })
            
            if activities:
                df = pd.DataFrame(activities)
                df['Original Duration'] = df['Original Duration'].fillna(1).astype(int)
                return {
                    'activities': df, 
                    'total_duration': df['Original Duration'].sum(), 
                    'total_cost': df['Budgeted Cost'].sum()
                }
            else:
                st.error("No activities found in XER file.")
                return None
    except Exception as e:
        st.error(f"Error parsing file: {str(e)}")
        return None

# ============================================================================
# EXTRACT PROJECT PARAMETERS FROM P6
# ============================================================================
def extract_project_parameters_from_p6(p6_result):
    """Extract project parameters from parsed P6 result"""
    if p6_result is None:
        return None
    
    if isinstance(p6_result, dict):
        activities_df = p6_result.get('activities')
        wbs_summary = p6_result.get('wbs_summary')
        total_cost_from_parse = p6_result.get('total_cost')
    else:
        activities_df = p6_result
        wbs_summary = None
        total_cost_from_parse = None
    
    if activities_df is None or activities_df.empty:
        return None
    
    # Get total cost
    if total_cost_from_parse and total_cost_from_parse > 0:
        total_project_cost = total_cost_from_parse
    else:
        if 'target_cost' in activities_df.columns:
            mask = activities_df['Activity ID'] != 'Indirect'
            total_project_cost = activities_df[mask]['target_cost'].sum()
        elif 'Budgeted Cost' in activities_df.columns:
            mask = activities_df['Activity ID'] != 'Indirect'
            total_project_cost = activities_df[mask]['Budgeted Cost'].sum()
        else:
            total_project_cost = 0
    
    # Get duration from Indirect activity
    total_duration = 1360
    
    indirect_activity = activities_df[
        (activities_df['Activity Name'].str.contains('Indirect', case=False, na=False)) |
        (activities_df['Activity ID'] == 'Indirect')
    ]
    
    if not indirect_activity.empty:
        if 'Original Duration' in indirect_activity.columns:
            duration = indirect_activity['Original Duration'].iloc[0]
            if duration > 0:
                total_duration = duration
        elif 'target_drtn_hr_cnt' in indirect_activity.columns:
            duration_hrs = indirect_activity['target_drtn_hr_cnt'].iloc[0]
            if duration_hrs > 0:
                total_duration = int(duration_hrs / 8)
    
    # Get indirect cost and rate
    indirect_cost = 0
    indirect_rate = 7540
    
    if not indirect_activity.empty:
        if 'target_cost' in indirect_activity.columns:
            indirect_cost = indirect_activity['target_cost'].iloc[0]
        elif 'Budgeted Cost' in indirect_activity.columns:
            indirect_cost = indirect_activity['Budgeted Cost'].iloc[0]
        
        if indirect_cost > 0 and total_duration > 0:
            indirect_rate = indirect_cost / total_duration
    
    # Activity statistics
    work_activities = activities_df[~activities_df['Activity ID'].isin(['Start', 'Finish', 'Indirect'])]
    
    activity_stats = {
        'total_activities': len(work_activities),
        'completed_activities': 0,
        'in_progress_activities': 0,
        'not_started_activities': len(work_activities),
        'avg_duration': work_activities['Original Duration'].mean() if 'Original Duration' in work_activities.columns else 0,
        'max_duration': work_activities['Original Duration'].max() if 'Original Duration' in work_activities.columns else 0,
        'min_duration': work_activities['Original Duration'].min() if 'Original Duration' in work_activities.columns else 0,
    }
    
    # Project name
    project_name = "Underground Metro Station"
    if wbs_summary and wbs_summary.get('project_name'):
        project_name = wbs_summary['project_name']
    elif 'WBS Name' in work_activities.columns:
        wbs_names = work_activities['WBS Name'].dropna().unique()
        if len(wbs_names) > 0:
            project_name = wbs_names[0]
    
    # Project location
    project_location = "Middle East"
    project_name_lower = project_name.lower()
    if 'dubai' in project_name_lower or 'uae' in project_name_lower:
        project_location = "United Arab Emirates"
    elif 'riyadh' in project_name_lower or 'saudi' in project_name_lower:
        project_location = "Saudi Arabia"
    elif 'doha' in project_name_lower or 'qatar' in project_name_lower:
        project_location = "Qatar"
    
    return {
        'project_name': project_name,
        'project_location': project_location,
        'project_type': "Metro/Transit Station",
        'total_duration': int(total_duration),
        'total_direct_cost': total_project_cost,
        'indirect_cost': indirect_cost,
        'indirect_rate': indirect_rate,
        'total_cost': total_project_cost,
        'critical_path_length': int(total_duration * 0.6),
        'activity_stats': activity_stats,
        'currency_detected': "USD"
    }

# ============================================================================
# EXTRACT RISKS FROM P6
# ============================================================================
def extract_risks_from_p6(activities_df):
    if activities_df is None or activities_df.empty:
        return None
    total_activities = len(activities_df)
    weather_activities = sum(1 for aid in activities_df['Activity ID'] if aid.startswith(('A300', 'A500', 'A700', 'A800')))
    weather_percentage = weather_activities / total_activities if total_activities > 0 else 0
    labor_activities = sum(1 for aid in activities_df['Activity ID'] if aid.startswith(('A600', 'A900')))
    labor_percentage = labor_activities / total_activities if total_activities > 0 else 0
    supply_activities = sum(1 for aid in activities_df['Activity ID'] if aid.startswith(('A200', 'A400', 'A500', 'A600', 'A700', 'A800', 'A900')))
    supply_percentage = supply_activities / total_activities if total_activities > 0 else 0
    risks_data = [
        {'Risk Name': 'Geopolitical Instability', 'Type': 'Threat', 
         'Probability (0-1)': round(0.25 + (supply_percentage * 0.3), 2), 
         'Impact (0-1)': round(0.15 + (supply_percentage * 0.2), 2), 
         'Distribution': 'Triangular',
         'Optimistic (0-1)': round(0.05 + (supply_percentage * 0.1), 2),
         'Most Likely (0-1)': round(0.15 + (supply_percentage * 0.2), 2),
         'Pessimistic (0-1)': round(0.30 + (supply_percentage * 0.3), 2),
         'Correlated With': 'Supply Chain Disruption', 'Correlation Strength': 0.7,
         'Response Strategy': 'Transfer', 'Response Plan': 'Political risk insurance, diversify suppliers'},
        {'Risk Name': 'Pandemic Disruption', 'Type': 'Threat', 'Probability (0-1)': 0.35, 'Impact (0-1)': 0.25,
         'Distribution': 'Triangular', 'Optimistic (0-1)': 0.10, 'Most Likely (0-1)': 0.25, 'Pessimistic (0-1)': 0.50,
         'Correlated With': '', 'Correlation Strength': 0.0, 'Response Strategy': 'Mitigate',
         'Response Plan': 'Health protocols, remote work capabilities'},
        {'Risk Name': 'Supply Chain Disruption', 'Type': 'Threat',
         'Probability (0-1)': round(0.20 + (supply_percentage * 0.3), 2),
         'Impact (0-1)': round(0.10 + (supply_percentage * 0.2), 2),
         'Distribution': 'Triangular',
         'Optimistic (0-1)': round(0.03 + (supply_percentage * 0.07), 2),
         'Most Likely (0-1)': round(0.10 + (supply_percentage * 0.2), 2),
         'Pessimistic (0-1)': round(0.25 + (supply_percentage * 0.3), 2),
         'Correlated With': 'Geopolitical Instability', 'Correlation Strength': 0.7,
         'Response Strategy': 'Mitigate', 'Response Plan': 'Dual sourcing, safety stock'},
        {'Risk Name': 'Labor Shortage', 'Type': 'Threat',
         'Probability (0-1)': round(0.15 + (labor_percentage * 0.4), 2),
         'Impact (0-1)': round(0.08 + (labor_percentage * 0.15), 2),
         'Distribution': 'Triangular',
         'Optimistic (0-1)': round(0.02 + (labor_percentage * 0.05), 2),
         'Most Likely (0-1)': round(0.08 + (labor_percentage * 0.15), 2),
         'Pessimistic (0-1)': round(0.20 + (labor_percentage * 0.25), 2),
         'Correlated With': '', 'Correlation Strength': 0.0, 'Response Strategy': 'Mitigate',
         'Response Plan': 'Recruitment plan, training programs'},
        {'Risk Name': 'Weather Events', 'Type': 'Threat',
         'Probability (0-1)': round(0.10 + (weather_percentage * 0.3), 2),
         'Impact (0-1)': round(0.05 + (weather_percentage * 0.12), 2),
         'Distribution': 'Triangular',
         'Optimistic (0-1)': round(0.01 + (weather_percentage * 0.03), 2),
         'Most Likely (0-1)': round(0.05 + (weather_percentage * 0.12), 2),
         'Pessimistic (0-1)': round(0.15 + (weather_percentage * 0.25), 2),
         'Correlated With': '', 'Correlation Strength': 0.0, 'Response Strategy': 'Accept',
         'Response Plan': 'Weather monitoring, schedule flexibility'}
    ]
    return pd.DataFrame(risks_data)

# ============================================================================
# SIDEBAR - PROJECT INFORMATION AND PARAMETERS
# ============================================================================
with st.sidebar:
    st.markdown("### PROJECT INFORMATION")
    
    current_project_name = st.session_state.auto_project_name if st.session_state.auto_project_name else "Underground Metro Station"
    current_project_location = st.session_state.auto_project_location if st.session_state.auto_project_location else "Middle East"
    
    project_name = st.text_input("Project Name", value=current_project_name, key="sidebar_project_name")
    project_location = st.text_input("Location", value=current_project_location, key="sidebar_project_location")
    project_manager = st.text_input("Project Manager", value="", key="sidebar_project_manager")
    
    detected_currency = "USD"
    for country, currency in country_currency_map.items():
        if country.lower() in project_location.lower():
            detected_currency = currency
            break
    
    st.markdown("---")
    st.markdown("### PROJECT PARAMETERS")
    
    if 'sidebar_version' not in st.session_state:
        st.session_state.sidebar_version = 0
    
    if st.session_state.auto_duration and st.session_state.auto_duration > 0:
        default_duration = st.session_state.auto_duration
    else:
        default_duration = 1360
    
    if st.session_state.auto_indirect_cost and st.session_state.auto_indirect_cost > 0:
        default_indirect_cost = st.session_state.auto_indirect_cost
    else:
        default_indirect_cost = 10254400
    
    if st.session_state.auto_total_cost and st.session_state.auto_total_cost > 0:
        default_total_cost = st.session_state.auto_total_cost
    else:
        default_total_cost = 61530799
    
    extracted_direct_cost = default_total_cost - default_indirect_cost
    
    if st.session_state.auto_indirect_rate and st.session_state.auto_indirect_rate > 0:
        default_indirect_rate = st.session_state.auto_indirect_rate
    else:
        default_indirect_rate = 7540
    
    # Show import status
    if st.session_state.p6_activities is not None:
        st.success(f"📊 P6 Imported: {default_duration} days | ${default_total_cost:,.0f}")
    else:
        st.info("💡 Import Primavera P6 Excel file above or edit parameters below")
    
    version = st.session_state.sidebar_version
    
    project_duration = st.number_input(
        "Baseline Duration (days)", 
        min_value=1, 
        value=int(default_duration), 
        step=10, 
        key=f"duration_input_{version}"
    )
    
    direct_cost = st.number_input(
        "Direct Cost ($)", 
        min_value=0, 
        value=int(extracted_direct_cost), 
        step=1000000, 
        format="%d", 
        key=f"direct_cost_input_{version}"
    )
    
    indirect_cost = st.number_input(
        "Indirect Cost ($)", 
        min_value=0, 
        value=int(default_indirect_cost), 
        step=1000000, 
        format="%d", 
        key=f"indirect_cost_input_{version}"
    )
    
    indirect_rate = st.number_input(
        "Indirect Cost Rate ($/day)", 
        min_value=0, 
        value=int(default_indirect_rate), 
        step=500, 
        key=f"indirect_rate_input_{version}"
    )
    
    total_cost = direct_cost + indirect_cost
    
    st.markdown("---")
    st.markdown("### COST SUMMARY")
    st.metric("Direct Cost", f"${direct_cost:,.0f}")
    st.metric("Indirect Cost", f"${indirect_cost:,.0f}")
    st.metric("Total Cost", f"${total_cost:,.0f}")
    st.markdown("---")
    st.markdown(f"**Project:** {project_name}")
    st.markdown(f"**Location:** {project_location}")
    st.markdown(f"**Currency:** {detected_currency}")
    if project_manager:
        st.markdown(f"**Manager:** {project_manager}")

    st.markdown("---")
    st.markdown("### SIMULATION SETTINGS")
    
    if 'uncertainty_value' not in st.session_state:
        st.session_state.uncertainty_value = 15
    
    activity_uncertainty_pct = st.slider(
        "Activity Duration Uncertainty (%)", 
        min_value=0, 
        max_value=50, 
        value=st.session_state.uncertainty_value, 
        step=5,
        help="Maximum % an activity duration can increase due to normal variance"
    )
    
    st.session_state.uncertainty_value = activity_uncertainty_pct
    
    st.markdown("---")
    st.markdown("### RISK MITIGATION LEVEL")
    
    if 'mitigation_level' not in st.session_state:
        st.session_state.mitigation_level = "Post-Mitigation"
    
    mitigation_level = st.radio(
        "Select Mitigation Level",
        ["Pre-Mitigation (Original Risks)", "Post-Mitigation (After Response Plans)"],
        index=0 if st.session_state.mitigation_level == "Pre-Mitigation" else 1,
        help="Pre-Mitigation: Original risk probabilities and impacts before response plans.\nPost-Mitigation: Reduced risks after applying response strategies."
    )
    
    st.session_state.mitigation_level = mitigation_level.replace(" (Original Risks)", "").replace(" (After Response Plans)", "")

# ============================================================================
# PRIMAVERA P6 IMPORT
# ============================================================================
st.markdown('<div class="card">', unsafe_allow_html=True)
st.markdown('<div class="card-title">📁 PRIMAVERA P6 INTEGRATION</div>', unsafe_allow_html=True)

uploaded_file = st.file_uploader("Import Primavera P6 Excel File", type=["xlsx", "xls"], key="p6_uploader")

if uploaded_file is not None:
    with st.spinner(f"Parsing {uploaded_file.name}..."):
        file_ext = uploaded_file.name.split('.')[-1].lower()
        file_content = uploaded_file.read()
        
        if file_ext in ['xlsx', 'xls']:
            p6_result = parse_excel_file(file_content)
            if p6_result is not None:
                p6_data = p6_result.get('activities')
                total_duration_from_p6 = p6_result.get('total_duration')
                total_cost_from_p6 = p6_result.get('total_cost')
                indirect_cost_from_p6 = p6_result.get('indirect_cost')
                project_name_from_p6 = p6_result.get('project_name')
                
                if p6_data is not None and len(p6_data) > 0:
                    st.session_state.p6_activities = p6_data
                    
                    if total_duration_from_p6 and total_duration_from_p6 > 0:
                        st.session_state.auto_duration = total_duration_from_p6
                    
                    if total_cost_from_p6 and total_cost_from_p6 > 0:
                        st.session_state.auto_total_cost = total_cost_from_p6
                    
                    if indirect_cost_from_p6 and indirect_cost_from_p6 > 0:
                        st.session_state.auto_indirect_cost = indirect_cost_from_p6
                    
                    if project_name_from_p6:
                        st.session_state.auto_project_name = project_name_from_p6
                    
                    st.session_state.sidebar_version = st.session_state.get('sidebar_version', 0) + 1
                    
                    st.toast(f"✅ Imported {len(p6_data)} activities", icon="📁")
                    
                    with st.expander("📊 Extracted Project Data", expanded=False):
                        st.markdown(f"""
                        - **Activities:** {len(p6_data)}
                        - **Duration:** {total_duration_from_p6} days
                        - **Total Cost:** ${total_cost_from_p6:,.0f}
                        - **Direct Cost:** ${(total_cost_from_p6 - indirect_cost_from_p6):,.0f}
                        - **Indirect Cost:** ${indirect_cost_from_p6:,.0f}
                        - **Project:** {project_name_from_p6}
                        """)
                else:
                    st.error("No activities found in Excel file")
            else:
                st.error("Failed to parse Excel file")
        else:
            st.error("Please upload an Excel file (.xlsx or .xls)")

st.markdown('</div>', unsafe_allow_html=True)

# Update sidebar button
col_btn1, col_btn2, col_btn3 = st.columns([1, 2, 1])
with col_btn2:
    if st.button("🔄 UPDATE SIDEBAR WITH IMPORTED VALUES", use_container_width=True, type="primary"):
        st.session_state.sidebar_version = st.session_state.get('sidebar_version', 0) + 1
        st.rerun()

# ============================================================================
# RISK-TO-ACTIVITY MAPPING DISPLAY
# ============================================================================
if st.session_state.p6_activities is not None:
    st.markdown('<div class="card">', unsafe_allow_html=True)
    st.markdown('<div class="card-title">🎯 RISK-TO-ACTIVITY MAPPING</div>', unsafe_allow_html=True)
    
    risk_mapping_config = {
        "Weather Delays": {
            "prefixes": ["A100", "A300", "A500"],
            "desc": "Outdoor activities affected by weather",
            "coeff": 0.60
        },
        "Labor Shortage": {
            "prefixes": ["A600", "A900"],
            "desc": "Labor-intensive activities",
            "coeff": 0.70
        },
        "Material Supply Delay": {
            "prefixes": ["A200", "A400", "A500", "A600", "A700"],
            "desc": "Material-dependent activities",
            "coeff": 0.75
        },
        "Ground Conditions": {
            "prefixes": ["A200", "A300", "A400"],
            "desc": "Underground/ground-related activities",
            "coeff": 0.70
        },
        "Equipment Breakdown": {
            "prefixes": ["A300", "A400"],
            "desc": "Equipment-intensive activities",
            "coeff": 0.65
        },
        "War / Geopolitical Instability": {
            "prefixes": ["A200", "A400", "A600"],
            "desc": "Critical material-intensive activities",
            "coeff": 0.50
        },
        "Pandemic Disruption": {
            "prefixes": ["A300", "A600", "A900"],
            "desc": "Labor-intensive on-site activities",
            "coeff": 0.35
        },
        "Supply Chain Volatility": {
            "prefixes": ["A200", "A400", "A500", "A600", "A700", "A800", "A900"],
            "desc": "Material-dependent activities",
            "coeff": 0.70
        },
        "Regulatory Changes": {
            "prefixes": ["A100", "A900"],
            "desc": "Approval and compliance activities",
            "coeff": 0.50
        },
        "Financial Volatility": {
            "prefixes": ["ALL"],
            "desc": "All activities (cost escalation)",
            "coeff": 0.40
        },
        "Design Changes": {
            "prefixes": ["A200", "A500", "A600", "A900"],
            "desc": "Design-dependent activities",
            "coeff": 0.60
        },
        "Safety Incident": {
            "prefixes": ["RANDOM_5%"],
            "desc": "Random 5% of activities (unpredictable)",
            "coeff": 0.30
        },
        "Extreme Event": {
            "prefixes": ["RANDOM_10%"],
            "desc": "Random 10% of activities (rare but severe)",
            "coeff": 0.50
        }
    }
    
    all_activity_ids = st.session_state.p6_activities['Activity ID'].tolist()
    
    mapping_summary = []
    for risk_name, risk_config in risk_mapping_config.items():
        affected_count = 0
        if risk_config["prefixes"] == ["ALL"]:
            affected_count = len(all_activity_ids)
        elif "RANDOM" in str(risk_config["prefixes"][0]):
            if "5%" in risk_config["prefixes"][0]:
                affected_count = int(len(all_activity_ids) * 0.05)
            else:
                affected_count = int(len(all_activity_ids) * 0.10)
        else:
            affected_ids_set = set()
            for prefix in risk_config["prefixes"]:
                matched = [aid for aid in all_activity_ids if aid.startswith(prefix)]
                affected_ids_set.update(matched)
            affected_count = len(affected_ids_set)
        
        mapping_summary.append({
            'Risk Factor': risk_name,
            'Affected Activities': affected_count,
            'Percentage of Total': f"{affected_count/len(all_activity_ids)*100:.1f}%",
            'Description': risk_config['desc']
        })
    
    mapping_df = pd.DataFrame(mapping_summary)
    st.dataframe(mapping_df, use_container_width=True, hide_index=True)
    
    with st.expander("🔍 View detailed activity mapping per risk"):
        selected_risk = st.selectbox("Select risk to see affected activities:", list(risk_mapping_config.keys()), key="risk_select_mapping_updated")
        
        if selected_risk:
            config = risk_mapping_config[selected_risk]
            affected_ids = []
            
            if config["prefixes"] == ["ALL"]:
                affected_ids = all_activity_ids
            elif "RANDOM" in str(config["prefixes"][0]):
                st.info(f"{selected_risk} affects approximately {len(all_activity_ids) * (5 if '5%' in config['prefixes'][0] else 10) // 100} random activities. Below are sample activities that could be affected:")
                sample_ids = all_activity_ids[:14] if '5%' in config['prefixes'][0] else all_activity_ids[:28]
                affected_ids = sample_ids
            else:
                for prefix in config["prefixes"]:
                    matched = [aid for aid in all_activity_ids if aid.startswith(prefix)]
                    affected_ids.extend(matched)
                affected_ids = list(set(affected_ids))
            
            if affected_ids:
                available_cols = st.session_state.p6_activities.columns.tolist()
                display_cols = ['Activity ID', 'Activity Name', 'Original Duration']
                
                if 'WBS Name' in available_cols:
                    display_cols.append('WBS Name')
                elif 'WBS Code' in available_cols:
                    display_cols.append('WBS Code')
                
                affected_df = st.session_state.p6_activities[st.session_state.p6_activities['Activity ID'].isin(affected_ids)][display_cols]
                st.dataframe(affected_df, use_container_width=True, hide_index=True)
                st.caption(f"Total: {len(affected_ids)} activities affected")
            else:
                st.info("No specific activities mapped for this risk")
    
    st.markdown('</div>', unsafe_allow_html=True)

# ============================================================================
# RISK REGISTER TABLE
# ============================================================================
st.markdown('<div class="card">', unsafe_allow_html=True)
st.markdown('<div class="card-title">RISK REGISTER</div>', unsafe_allow_html=True)

risk_columns = ['Risk Name', 'Type', 'Probability (0-1)', 'Impact (0-1)', 'Distribution', 
                'Optimistic (0-1)', 'Most Likely (0-1)', 'Pessimistic (0-1)', 
                'Correlated With', 'Correlation Strength', 'Response Strategy', 'Response Plan']

default_risk_factors = pd.DataFrame({
    'Risk Name': ['Geopolitical Instability', 'Pandemic Disruption', 'Supply Chain Disruption', 'Labor Shortage', 'Weather Events'],
    'Type': ['Threat', 'Threat', 'Threat', 'Threat', 'Threat'],
    'Probability (0-1)': [0.40, 0.35, 0.30, 0.25, 0.15],
    'Impact (0-1)': [0.25, 0.20, 0.15, 0.12, 0.08],
    'Distribution': ['Triangular', 'Triangular', 'Triangular', 'Triangular', 'Triangular'],
    'Optimistic (0-1)': [0.10, 0.05, 0.05, 0.05, 0.02],
    'Most Likely (0-1)': [0.25, 0.20, 0.15, 0.12, 0.08],
    'Pessimistic (0-1)': [0.40, 0.35, 0.30, 0.25, 0.15],
    'Correlated With': ['', 'Geopolitical Instability', 'Geopolitical Instability', '', ''],
    'Correlation Strength': [0.0, 0.8, 0.7, 0.0, 0.0],
    'Response Strategy': ['Transfer', 'Mitigate', 'Mitigate', 'Mitigate', 'Accept'],
    'Response Plan': ['Political risk insurance, diversify suppliers', 'Health protocols, remote work plans', 'Dual sourcing, inventory buffer', 'Recruitment plan, training', 'Weather monitoring, schedule flexibility']
})

if 'Probability (0-1)' not in st.session_state.risk_factors.columns:
    st.session_state.risk_factors = default_risk_factors.copy()
else:
    if pd.isna(st.session_state.risk_factors['Probability (0-1)'].iloc[0]):
        st.session_state.risk_factors = default_risk_factors.copy()

for col in risk_columns:
    if col not in st.session_state.risk_factors.columns:
        if col in ['Correlated With', 'Response Strategy', 'Response Plan']:
            st.session_state.risk_factors[col] = ''
        elif col == 'Correlation Strength':
            st.session_state.risk_factors[col] = 0.0
        else:
            st.session_state.risk_factors[col] = None

risk_name_options = [''] + st.session_state.risk_factors['Risk Name'].tolist()

edited_df = st.data_editor(
    st.session_state.risk_factors[risk_columns],
    use_container_width=True,
    hide_index=True,
    column_config={
        "Risk Name": st.column_config.TextColumn("Risk Name", width="medium", required=True),
        "Type": st.column_config.SelectboxColumn("Type", options=["Threat", "Opportunity"], required=True),
        "Probability (0-1)": st.column_config.NumberColumn("Probability (0-1)", min_value=0.0, max_value=1.0, step=0.01, format="%.2f"),
        "Impact (0-1)": st.column_config.NumberColumn("Impact (0-1)", min_value=0.0, max_value=1.0, step=0.01, format="%.2f"),
        "Distribution": st.column_config.SelectboxColumn("Distribution", options=["Triangular", "Uniform", "Beta (PERT)", "Normal"], required=True),
        "Optimistic (0-1)": st.column_config.NumberColumn("Optimistic (0-1)", min_value=0.0, max_value=1.0, step=0.01, format="%.2f"),
        "Most Likely (0-1)": st.column_config.NumberColumn("Most Likely (0-1)", min_value=0.0, max_value=1.0, step=0.01, format="%.2f"),
        "Pessimistic (0-1)": st.column_config.NumberColumn("Pessimistic (0-1)", min_value=0.0, max_value=1.0, step=0.01, format="%.2f"),
        "Correlated With": st.column_config.SelectboxColumn("Correlated With", options=risk_name_options, required=False),
        "Correlation Strength": st.column_config.NumberColumn("Correlation Strength", min_value=0.0, max_value=1.0, step=0.05, format="%.2f"),
        "Response Strategy": st.column_config.SelectboxColumn("Response Strategy", options=["Avoid", "Mitigate", "Transfer", "Accept"]),
        "Response Plan": st.column_config.TextColumn("Response Plan", width="large"),
    }
)

validation_errors = validate_risk_input(edited_df)
if validation_errors:
    st.error(f"Input validation errors: {', '.join(validation_errors)}")
    st.stop()

if not edited_df.equals(st.session_state.risk_factors[risk_columns]):
    for col in risk_columns:
        if col in edited_df.columns:
            st.session_state.risk_factors[col] = edited_df[col].values
    st.session_state.ai_trained = False
    st.rerun()

col1, col2, col3, col4 = st.columns([2, 1, 1, 1]) 
with col2:
    if st.button("ADD RISK", use_container_width=True):
        new_row = pd.DataFrame({
            'Risk Name': ['New Risk'], 'Type': ['Threat'], 
            'Probability (0-1)': [0.10], 'Impact (0-1)': [0.10], 
            'Distribution': ['Triangular'], 'Optimistic (0-1)': [0.05], 
            'Most Likely (0-1)': [0.10], 'Pessimistic (0-1)': [0.20], 
            'Correlated With': [''], 'Correlation Strength': [0.0], 
            'Response Strategy': ['Accept'], 'Response Plan': ['']
        })
        st.session_state.risk_factors = pd.concat([st.session_state.risk_factors, new_row], ignore_index=True)
        st.session_state.ai_trained = False
        st.rerun()

with col3:
    if st.button("DELETE LAST", use_container_width=True):
        if len(st.session_state.risk_factors) > 1:
            st.session_state.risk_factors = st.session_state.risk_factors.iloc[:-1]
            st.session_state.ai_trained = False
            st.rerun()

with col4:
    if st.button("🔄 AUTO-POPULATE", use_container_width=True):
        if st.session_state.p6_activities is not None:
            auto_risks = extract_risks_from_p6(st.session_state.p6_activities)
            if auto_risks is not None:
                st.session_state.risk_factors = auto_risks
                st.session_state.ai_trained = False
                st.success("Risk register auto-populated from schedule!")
                st.rerun()
        else:
            st.warning("Please import a Primavera P6 file first")

st.markdown('</div>', unsafe_allow_html=True)

# ============================================================================
# AI RISK PREDICTION
# ============================================================================
st.markdown('<div class="card">', unsafe_allow_html=True)
st.markdown('<div class="card-title">🤖 Artificial Intelligence Risk Forecast</div>', unsafe_allow_html=True)

col_ai1, col_ai2 = st.columns([1, 2])

def run_ai_prediction_with_lock():
    if st.session_state.operation_in_progress:
        st.warning("Another operation is in progress. Please wait...")
        return
    st.session_state.operation_in_progress = True
    try:
        with st.spinner("Analyzing risk factors..."):
            current_risks = st.session_state.risk_factors['Probability (0-1)'].values.reshape(1, -1)
            X_scaled = st.session_state.ai_scaler.transform(current_risks)
            prediction = st.session_state.ai_model.predict(X_scaled)[0]
            st.session_state.ai_prediction = prediction
            importance_df = pd.DataFrame({
                'Risk Factor': st.session_state.risk_factors['Risk Name'], 
                'Influence (%)': st.session_state.ai_model.feature_importances_ * 100
            }).sort_values('Influence (%)', ascending=False)
            st.session_state.ai_importance = importance_df
    finally:
        st.session_state.operation_in_progress = False
        st.rerun()

with col_ai1:
    if st.button("Train AI Model", use_container_width=True, key="train_ai_btn"):
        with st.status("Training AI Model...", expanded=True) as status:
            status.update(label="Generating training data...", state="running")
            time.sleep(0.5)
            status.update(label="Training Random Forest algorithm...", state="running")
            model, scaler = train_risk_prediction_model(st.session_state.risk_factors)
            st.session_state.ai_model = model
            st.session_state.ai_scaler = scaler
            st.session_state.ai_trained = True
            status.update(label="Model training complete!", state="complete")
            st.success("AI model ready")
    if st.session_state.get('ai_trained', False):
        if st.button("Predict Delay", use_container_width=True, key="predict_ai_btn"):
            run_ai_prediction_with_lock()

with col_ai2:
    if st.session_state.get('ai_trained', False) and st.session_state.get('ai_prediction') is not None:
        pred = st.session_state.ai_prediction
        if pred >= 0.7:
            st.error(f"**{pred:.1%}** - High delay risk. Immediate mitigation recommended.")
        elif pred >= 0.4:
            st.warning(f"**{pred:.1%}** - Moderate delay risk. Monitor risk factors.")
        else:
            st.success(f"**{pred:.1%}** - Low delay risk. Continue current management.")
        st.markdown("#### Primary Risk Drivers")
        top_risks = st.session_state.ai_importance.head(5).copy()
        top_risks['Influence (%)'] = top_risks['Influence (%)'].round(1)
        st.dataframe(top_risks, use_container_width=True, hide_index=True)
    elif st.session_state.get('ai_trained', False):
        st.info("Click 'Predict Delay' to analyze current risks.")
    else:
        st.info("Click 'Train AI Model' to initialize the prediction system.")

st.markdown('</div>', unsafe_allow_html=True)

# ============================================================================
# RISK RESPONSE PLANS
# ============================================================================
for idx, row in st.session_state.risk_factors.iterrows():
    risk_name = row['Risk Name']
    prob = row['Probability (0-1)']
    impact = normalize_impact(row, direct_cost, project_duration)
    risk_score = prob * impact
    
    with st.expander(f"{risk_name} | Score: {risk_score:.0%} | {row.get('Response Strategy', 'Accept')}"):
        col_a, col_b = st.columns(2)
        with col_a:
            st.markdown(f"**Risk Score:** {risk_score:.0%}")
            st.markdown(f"**Probability:** {prob:.0%}")
            st.markdown(f"**Impact (Normalized):** {impact:.0%}")
        with col_b:
            st.markdown(f"**Response Strategy:** {row.get('Response Strategy', 'Accept')}")
            st.markdown(f"**Response Plan:** {row.get('Response Plan', 'Not defined')}")
        
        valid_strategies = ["Avoid", "Mitigate", "Transfer", "Accept"]
        current_strategy = row.get('Response Strategy', 'Accept')
        
        if current_strategy not in valid_strategies:
            current_strategy = "Mitigate"
        
        new_strategy = st.selectbox(
            "Update Response Strategy",
            valid_strategies,
            index=valid_strategies.index(current_strategy),
            key=f"resp_strat_{idx}"
        )
        new_plan = st.text_area(
            "Update Action Plan",
            value=row.get('Response Plan', ''),
            height=80,
            key=f"resp_plan_{idx}"
        )
        
        if st.button(f"Update {risk_name}", key=f"update_resp_{idx}"):
            st.session_state.risk_factors.at[idx, 'Response Strategy'] = new_strategy
            st.session_state.risk_factors.at[idx, 'Response Plan'] = new_plan
            st.success(f"Updated {risk_name}")
            st.rerun()

st.markdown('</div>', unsafe_allow_html=True)

# ============================================================================
# RISK MATRIX THRESHOLDS
# ============================================================================
st.markdown('<div class="card">', unsafe_allow_html=True)
st.markdown('<div class="card-title">RISK MATRIX THRESHOLDS</div>', unsafe_allow_html=True)

col1, col2 = st.columns(2)
with col1:
    st.markdown("**Probability Thresholds**")
    high_prob = st.slider("High Probability (≥)", min_value=0.0, max_value=1.0, value=st.session_state.risk_thresholds['high_prob'], step=0.05)
    medium_prob = st.slider("Medium Probability (≥)", min_value=0.0, max_value=1.0, value=st.session_state.risk_thresholds['medium_prob'], step=0.05)
    low_prob = st.slider("Low Probability (≥)", min_value=0.0, max_value=1.0, value=st.session_state.risk_thresholds['low_prob'], step=0.05)

with col2:
    st.markdown("**Impact Thresholds (Normalized 0-1 Scale)**")
    high_impact = st.slider("High Impact (≥)", min_value=0.0, max_value=1.0, value=st.session_state.risk_thresholds['high_impact'], step=0.05)
    medium_impact = st.slider("Medium Impact (≥)", min_value=0.0, max_value=1.0, value=st.session_state.risk_thresholds['medium_impact'], step=0.05)
    low_impact = st.slider("Low Impact (≥)", min_value=0.0, max_value=1.0, value=st.session_state.risk_thresholds['low_impact'], step=0.05)

st.session_state.risk_thresholds = {
    'high_prob': high_prob, 'medium_prob': medium_prob, 'low_prob': low_prob,
    'high_impact': high_impact, 'medium_impact': medium_impact, 'low_impact': low_impact
}

st.markdown('</div>', unsafe_allow_html=True)

# ============================================================================
# RISK CORRELATION MATRIX
# ============================================================================
# ============================================================================
# RISK CORRELATION MATRIX
# ============================================================================
st.markdown('<div class="card">', unsafe_allow_html=True)
st.markdown('<div class="card-title">🔗 Risk Correlation Matrix</div>', unsafe_allow_html=True)

risk_scores = []
for _, row in st.session_state.risk_factors.iterrows():
    norm_impact = normalize_impact(row, direct_cost, project_duration)
    risk_scores.append(row['Probability (0-1)'] * norm_impact)

ordered_risks = st.session_state.risk_factors.copy()
ordered_risks['Risk Score'] = risk_scores
ordered_risks = ordered_risks.sort_values('Risk Score', ascending=False)

risk_names_ordered = ordered_risks['Risk Name'].tolist()
n_risks = len(risk_names_ordered)
corr_matrix = np.eye(n_risks)

for i, row in st.session_state.risk_factors.iterrows():
    correlated_with = row.get('Correlated With', '')
    corr_strength = row.get('Correlation Strength', 0.0)
    if correlated_with and correlated_with in risk_names_ordered:
        current_name = row['Risk Name']
        if current_name in risk_names_ordered and current_name != correlated_with:
            i_ordered = risk_names_ordered.index(current_name)
            j_ordered = risk_names_ordered.index(correlated_with)
            if i_ordered != j_ordered:
                corr_matrix[i_ordered, j_ordered] = corr_strength
                corr_matrix[j_ordered, i_ordered] = corr_strength

corr_matrix = make_positive_definite(corr_matrix)

fig_corr = px.imshow(
    corr_matrix,
    labels=dict(x="", y="", color=""),
    x=risk_names_ordered,
    y=risk_names_ordered,
    color_continuous_scale=['#22c55e', '#eab308', '#ef4444'],
    title="Risk Correlation Matrix",
    aspect="auto",
    zmin=0,
    zmax=1,
    text_auto='.2f'
)

# UPDATE LAYOUT WITH EXPLICIT FONT TO REMOVE ITALIC
fig_corr.update_layout(
    height=500,
    paper_bgcolor='#1e293b',
    plot_bgcolor='#1e293b',
    font_color='#cbd5e1',
    font=dict(family="Arial, sans-serif", size=12, color="#cbd5e1", weight="normal")
)

st.plotly_chart(fig_corr, use_container_width=True, config=chart_config)
st.markdown('</div>', unsafe_allow_html=True)

# ============================================================================
# EXTERNAL DATA INTEGRATION
# ============================================================================
st.markdown('<div class="card">', unsafe_allow_html=True)
st.markdown('<div class="card-title">🌍 EXTERNAL DATA INTEGRATION</div>', unsafe_allow_html=True)

col_api1, col_api2 = st.columns(2)

with col_api1:
    st.markdown("#### Weather Conditions")
    selected_country = st.selectbox("Country", country_options, help="Select project country")
    default_city = country_city_map.get(selected_country, selected_country)
    city_input = st.text_input("City", value=default_city, help="Enter city name")
    location_display = f"{city_input}, {selected_country}" if city_input else selected_country
    
    if st.button("Fetch Weather Data", use_container_width=True, key="fetch_weather"):
        with st.spinner(f"Retrieving weather for {location_display}..."):
            weather_data = get_free_weather(city_input, selected_country)
            if weather_data:
                st.session_state.weather_data = weather_data
                if weather_data.get('is_simulated'):
                    st.info("Using simulated weather data (API unavailable)")
                else:
                    st.success("Weather data retrieved")
            else:
                st.warning("Could not retrieve weather. Using simulated data.")
                st.session_state.weather_data = simulate_weather_data(location_display)
    
    if st.session_state.weather_data:
        w = st.session_state.weather_data
        current = w.get('current_weather', {})
        temp = current.get('temperature', 25)
        st.markdown(f"**Location:** {location_display}")
        st.markdown(f"**Temperature:** {temp:.1f}°C")
        st.markdown(f"**Wind Speed:** {current.get('windspeed', 0):.1f} km/h")
        if w.get('is_simulated'):
            st.caption("⚠️ Using simulated weather data")
        if temp > 40:
            st.warning("High temperature alert. Consider heat-related productivity adjustments.")

with col_api2:
    st.markdown("#### Commodity Market Data")
    
    if 'commodity_data' not in st.session_state:
        st.session_state.commodity_data = None
    
    if st.button("Fetch Commodity Prices", use_container_width=True, key="fetch_commodity"):
        with st.spinner(f"Retrieving commodity prices for {selected_country}..."):
            st.session_state.commodity_data = get_commodity_prices(selected_country)
            st.success(f"Commodity prices updated for {selected_country}")
    
    if st.session_state.commodity_data is not None:
        c = st.session_state.commodity_data
        st.markdown(f"**Steel:** {c['currency']}{c['steel']:.0f}/ton")
        st.markdown(f"**Cement:** {c['currency']}{c['cement']:.0f}/m³")
        st.markdown(f"**Fuel:** {c['currency']}{c['fuel']:.0f}/bbl")
        st.caption(f"Prices in {c['currency_code']} - Local market rates for {selected_country}")
    else:
        st.info("Click 'Fetch Commodity Prices' to see country-specific rates")

st.caption("External data integration enhances risk assessment with real-world conditions.")
st.markdown('</div>', unsafe_allow_html=True)

# ============================================================================
# PMI-RMP RISK MATRIX
# ============================================================================
st.markdown('<div class="card">', unsafe_allow_html=True)
st.markdown('<div class="card-title">📊 PMI-RMP RISK MATRIX</div>', unsafe_allow_html=True)

risk_levels = []
normalized_impacts = []

for _, row in st.session_state.risk_factors.iterrows():
    prob = row['Probability (0-1)']
    normalized_impact = normalize_impact(row, direct_cost, project_duration)
    normalized_impacts.append(normalized_impact)
    risk_score_value = prob * normalized_impact
    
    if risk_score_value >= 0.8:
        level = "🔴 CRITICAL"
    elif risk_score_value >= 0.7:
        level = "🔴 EXTREME"
    elif risk_score_value >= 0.5:
        level = "🟠 HIGH"
    elif risk_score_value >= 0.3:
        level = "🟡 MEDIUM"
    else:
        level = "🟢 LOW"
    
    risk_levels.append({
        'Risk': row['Risk Name'],
        'Probability': f"{prob:.0%}",
        'Impact': f"{normalized_impact:.0%}",
        'Score': f"{risk_score_value:.0%}",
        'Level': level
    })

st.dataframe(pd.DataFrame(risk_levels), use_container_width=True, hide_index=True)

matrix_data = []
cell_risk_counts = [[0 for _ in range(5)] for _ in range(5)]
impact_labels = ['Very Low', 'Low', 'Medium', 'High', 'Very High']
prob_labels = ['Very High', 'High', 'Medium', 'Low', 'Very Low']

for idx, (_, risk_row) in enumerate(st.session_state.risk_factors.iterrows()):
    prob = risk_row['Probability (0-1)']
    norm_impact = normalized_impacts[idx]
    
    if prob >= 0.8:
        prob_cell = 0
    elif prob >= 0.6:
        prob_cell = 1
    elif prob >= 0.4:
        prob_cell = 2
    elif prob >= 0.2:
        prob_cell = 3
    else:
        prob_cell = 4
    
    if norm_impact >= 0.8:
        impact_cell = 4
    elif norm_impact >= 0.6:
        impact_cell = 3
    elif norm_impact >= 0.4:
        impact_cell = 2
    elif norm_impact >= 0.2:
        impact_cell = 1
    else:
        impact_cell = 0
    
    cell_risk_counts[prob_cell][impact_cell] += 1

for i in range(5):
    row_data = []
    for j in range(5):
        risk_score = (1 - i * 0.2) * (j * 0.2 + 0.1)
        row_data.append(risk_score)
    matrix_data.append(row_data)

fig_matrix = px.imshow(
    matrix_data,
    labels=dict(x="Impact", y="Probability", color="Score"),
    x=impact_labels,
    y=prob_labels,
    color_continuous_scale=['#22c55e', '#eab308', '#f97316', '#dc2626', '#991b1b'],
    title="Probability-Impact Matrix",
    aspect="auto",
    zmin=0,
    zmax=1
)

for i in range(5):
    for j in range(5):
        if cell_risk_counts[i][j] > 0:
            fig_matrix.add_annotation(x=j, y=i, text=f"<b>{cell_risk_counts[i][j]}</b>", showarrow=False, font=dict(size=14, color="white", weight="bold"))

fig_matrix.update_layout(height=500, paper_bgcolor='#1e293b', plot_bgcolor='#1e293b', font_color='#cbd5e1')
st.plotly_chart(fig_matrix, use_container_width=True, config=chart_config)
st.markdown('</div>', unsafe_allow_html=True)

# ============================================================================
# EXPECTED MONETARY VALUE (EMV) ANALYSIS
# ============================================================================
st.markdown('<div class="card">', unsafe_allow_html=True)
st.markdown('<div class="card-title">💰 EXPECTED MONETARY VALUE (EMV) ANALYSIS</div>', unsafe_allow_html=True)

emv_data = []
total_emv_threat = 0
total_emv_opportunity = 0

for _, row in st.session_state.risk_factors.iterrows():
    prob = row['Probability (0-1)']
    emv, impact_amount = calculate_correct_emv(row, direct_cost, indirect_rate, project_duration)
    
    if row['Type'] == 'Threat':
        total_emv_threat += emv
        emv_sign = "-"
    else:
        total_emv_opportunity += emv
        emv_sign = "+"
    
    emv_data.append({
        'Risk': row['Risk Name'], 
        'Type': row['Type'], 
        'Probability': f"{prob:.0%}", 
        'EMV': f"{emv_sign}${abs(emv):,.0f}",
        'Potential Impact': f"${impact_amount:,.0f}"
    })

net_emv = total_emv_opportunity - total_emv_threat
st.dataframe(pd.DataFrame(emv_data), use_container_width=True, hide_index=True)

col1, col2, col3 = st.columns(3)
with col1:
    st.metric("Total Threat EMV", f"-${total_emv_threat:,.0f}")
with col2:
    st.metric("Total Opportunity EMV", f"+${total_emv_opportunity:,.0f}")
with col3:
    st.metric("NET EMV", f"${net_emv:,.0f}")

if net_emv < 0:
    st.warning(f"Overall risk exposure is negative (threats outweigh opportunities by ${abs(net_emv):,.0f})")
else:
    st.success(f"Overall risk exposure is positive (opportunities outweigh threats by ${net_emv:,.0f})")
st.markdown('</div>', unsafe_allow_html=True)

# ============================================================================
# DECISION TREE ANALYSIS
# ============================================================================
st.markdown('<div class="card">', unsafe_allow_html=True)
st.markdown('<div class="card-title">🌳 DECISION TREE ANALYSIS</div>', unsafe_allow_html=True)

st.markdown("""
**Use Cases:**
- **Investment Decision**: Compare multiple investment options with different costs and returns
- **Vendor Selection**: Choose between vendors with different pricing and reliability
- **Risk Response**: Decide between different mitigation strategies
""")

decision_type = st.radio(
    "Select Decision Type",
    ["Investment/Project Selection (with initial cost)", "Vendor/Contractor Selection (no initial cost)", "Custom Decision"],
    horizontal=True
)

col1, col2 = st.columns(2)
with col1:
    decision_name = st.text_input("Decision Name", placeholder="e.g., Choose Construction Method", value="MAIN DECISION")
with col2:
    num_alternatives = st.number_input("Number of Alternatives", min_value=2, max_value=5, value=2, step=1)

alternatives = []
for i in range(num_alternatives):
    with st.expander(f"Alternative {i+1}", expanded=False):
        alt_name = st.text_input(f"Alternative Name", key=f"alt_name_{i}", placeholder=f"Option {i+1}", value=f"Option {i+1}")
        
        if decision_type == "Investment/Project Selection (with initial cost)":
            initial_investment = st.number_input(f"Initial Investment", value=0.0, step=1000000.0, key=f"init_inv_{i}")
        else:
            initial_investment = 0
        
        num_outcomes = st.number_input(f"Number of possible outcomes", min_value=1, max_value=4, value=2, key=f"num_outcomes_{i}")
        alt_outcomes = []
        
        for j in range(num_outcomes):
            st.markdown(f"**Outcome {j+1}**")
            col_a, col_b, col_c = st.columns([2, 1, 1])
            with col_a:
                outcome_name = st.text_input(f"Outcome Name", key=f"outcome_name_{i}_{j}", placeholder="e.g., Success", value=f"Outcome {j+1}")
            with col_b:
                outcome_prob = st.number_input(f"Probability", min_value=0.0, max_value=1.0, value=0.5, step=0.05, key=f"outcome_prob_{i}_{j}")
            with col_c:
                outcome_value = st.number_input(f"Value", value=10000000.0, step=1000000.0, key=f"outcome_value_{i}_{j}")
            alt_outcomes.append({'name': outcome_name, 'Probability': outcome_prob, 'value': outcome_value})
        
        total_prob = sum(o['Probability'] for o in alt_outcomes)
        if abs(total_prob - 1.0) > 0.01:
            st.warning(f"Total probability is {total_prob:.0%}. Should sum to 100%")
        
        gross_emv = sum(o['Probability'] * o['value'] for o in alt_outcomes)
        net_emv_alt = gross_emv - initial_investment
        
        alternatives.append({
            'name': alt_name,
            'initial_investment': initial_investment,
            'outcomes': alt_outcomes,
            'gross_emv': gross_emv,
            'net_emv': net_emv_alt
        })

def run_decision_tree_with_lock():
    if st.session_state.operation_in_progress:
        st.warning("Another operation is in progress. Please wait...")
        return
    st.session_state.operation_in_progress = True
    
    try:
        best_alt = max(alternatives, key=lambda x: x['net_emv'])
        
        results_data = []
        for alt in alternatives:
            results_data.append({
                'Alternative': alt['name'],
                'Initial Investment': f"{alt['initial_investment']:,.0f}",
                'Gross EMV': f"{alt['gross_emv']:,.0f}",
                'Net EMV': f"{alt['net_emv']:,.0f}",
                'Recommendation': "✅ BEST" if alt['name'] == best_alt['name'] else ""
            })
        
        st.dataframe(pd.DataFrame(results_data), use_container_width=True, hide_index=True)
        
        if best_alt['net_emv'] < 0:
            st.error(f"**Recommended Decision: {best_alt['name']}**")
            st.caption(f"Expected Net Value: {best_alt['net_emv']:,.0f} (Negative - Consider alternatives)")
        else:
            st.success(f"**Recommended Decision: {best_alt['name']}**")
            st.caption(f"Expected Net Value: {best_alt['net_emv']:,.0f}")
        
        fig = go.Figure()
        fig.add_trace(go.Scatter(
            x=[0.5], y=[1],
            mode='text',
            text=[f"<b>{decision_name}</b>"],
            textposition="middle center",
            textfont=dict(size=14, color='#f1f5f9', family='Arial Black'),
            showlegend=False
        ))
        
        alt_x_positions = np.linspace(0.1, 0.9, len(alternatives))
        
        for idx, alt in enumerate(alternatives):
            x_pos = alt_x_positions[idx]
            fig.add_trace(go.Scatter(
                x=[x_pos], y=[0.65],
                mode='text',
                text=[f"<b>{alt['name']}</b><br>${alt['net_emv']:,.0f}"],
                textposition="middle center",
                textfont=dict(size=11, color='#f97316'),
                showlegend=False
            ))
            
            fig.add_annotation(
                x=x_pos, y=0.72,
                ax=0.5, ay=0.92,
                showarrow=True,
                arrowhead=2,
                arrowsize=1.5,
                arrowwidth=2,
                arrowcolor='#3b82f6'
            )
        
        fig.update_layout(
            title="<b>Decision Tree Visualization</b>",
            height=550,
            paper_bgcolor='#1e293b',
            plot_bgcolor='#1e293b',
            font_color='#cbd5e1',
            showlegend=False,
            xaxis=dict(showgrid=False, zeroline=False, visible=False, range=[-0.05, 1.05]),
            yaxis=dict(showgrid=False, zeroline=False, visible=False, range=[0.15, 1.05]),
        )
        
        st.plotly_chart(fig, use_container_width=True, config=chart_config)
        
    finally:
        st.session_state.operation_in_progress = False

if st.button("CALCULATE DECISION TREE", use_container_width=True, type="primary"):
    run_decision_tree_with_lock()

st.markdown('</div>', unsafe_allow_html=True)

# ============================================================================
# MONTE CARLO SIMULATION SETTINGS
# ============================================================================
st.markdown('<div class="card">', unsafe_allow_html=True)
st.markdown('<div class="card-title">🎲 MONTE CARLO SIMULATION SETTINGS</div>', unsafe_allow_html=True)

col1, col2 = st.columns(2)
with col1:
    iterations = st.number_input("Number of Iterations", min_value=1000, value=10000, step=1000)
with col2:
    percentile_input = st.text_input("Custom Percentiles", value="50,80,90")
    try:
        user_percentiles = [int(p.strip()) for p in percentile_input.split(",")]
    except:
        user_percentiles = [50, 80, 90]

def run_simulation_with_lock():
    if st.session_state.operation_in_progress:
        st.warning("Another operation is in progress. Please wait...")
        return
    
    param_errors = validate_project_parameters(project_duration, direct_cost, indirect_rate)
    if param_errors:
        st.error(f"Parameter validation failed: {', '.join(param_errors)}")
        return
    
    st.session_state.operation_in_progress = True
    
    progress_container = st.container()
    with progress_container:
        progress_bar = st.progress(0)
        status_text = st.empty()
    
    start_time = time.time()
    
    throttled_callback = ThrottledProgressCallback(progress_bar, status_text, iterations, update_interval_pct=0.05)
    
    try:
        durations, costs, risk_contributions, iteration_details_df = run_monte_carlo_full_optimized(
            st.session_state.risk_factors, 
            project_duration, 
            total_cost,
            indirect_rate, 
            iterations, 
            throttled_callback,
            activities_df=st.session_state.p6_activities,
            pra_data=None
        )
        
        if durations is None:
            st.error("Simulation failed due to input validation errors")
            return
        
        st.session_state.simulation_results = {
            'durations': durations, 
            'costs': costs, 
            'risk_contributions': risk_contributions
        }
        st.session_state.simulation_details = iteration_details_df
        st.session_state.simulation_run_flag = True
        
        progress_container.empty()
        
        elapsed_time = time.time() - start_time
        st.success(f"✅ Simulation completed! {iterations:,} iterations processed in {elapsed_time:.1f} seconds")
        
    except Exception as e:
        st.error(f"Simulation error: {str(e)}")
        progress_container.empty()
    finally:
        st.session_state.operation_in_progress = False

run_button = st.button("🚀 RUN MONTE CARLO SIMULATION", type="primary", use_container_width=True)

if run_button:
    run_simulation_with_lock()

st.markdown('</div>', unsafe_allow_html=True)

# ============================================================================
# DISPLAY SIMULATION RESULTS
# ============================================================================
if st.session_state.simulation_run_flag and st.session_state.simulation_results is not None:
    durations = st.session_state.simulation_results['durations']
    costs = st.session_state.simulation_results['costs']
    risk_contributions = st.session_state.simulation_results['risk_contributions']
    
    percentiles = {}
    for p in user_percentiles:
        if 0 <= p <= 100:
            percentiles[f'P{p}'] = np.percentile(durations, p)
    
    cost_percentiles = {}
    for p in user_percentiles:
        if 0 <= p <= 100:
            cost_percentiles[f'P{p}'] = np.percentile(costs, p)
    
    percentile_colors = ['#ef4444', '#f97316', '#eab308', '#10b981', '#06b6d4', '#8b5cf6', '#ec4899', '#f43f5e', '#14b8a6']
    
    # Statistical Confidence Intervals
    with st.expander("📊 Statistical Confidence Intervals", expanded=False):
        duration_ci = calculate_confidence_intervals(durations)
        cost_ci = calculate_confidence_intervals(costs)
        
        col_ci1, col_ci2 = st.columns(2)
        
        with col_ci1:
            st.markdown("#### Duration Confidence Intervals")
            duration_ci_data = []
            for key, value in duration_ci.items():
                if isinstance(value, dict):
                    duration_ci_data.append({
                        'Confidence Level': key,
                        'Lower Bound': f"{value['lower']:.0f} days",
                        'Upper Bound': f"{value['upper']:.0f} days",
                        'Range': f"{value['range']:.0f} days"
                    })
            duration_ci_data.append({'Confidence Level': 'Mean', 'Lower Bound': '', 'Upper Bound': f"{duration_ci['mean']:.0f} days", 'Range': ''})
            duration_ci_data.append({'Confidence Level': 'Std Dev', 'Lower Bound': '', 'Upper Bound': f"{duration_ci['std_error']:.1f} days", 'Range': ''})
            st.dataframe(pd.DataFrame(duration_ci_data), use_container_width=True, hide_index=True)
        
        with col_ci2:
            st.markdown("#### Cost Confidence Intervals")
            cost_ci_data = []
            for key, value in cost_ci.items():
                if isinstance(value, dict):
                    cost_ci_data.append({
                        'Confidence Level': key,
                        'Lower Bound': f"${value['lower']:,.0f}",
                        'Upper Bound': f"${value['upper']:,.0f}",
                        'Range': f"${value['range']:,.0f}"
                    })
            cost_ci_data.append({'Confidence Level': 'Mean', 'Lower Bound': '', 'Upper Bound': f"${cost_ci['mean']:,.0f}", 'Range': ''})
            cost_ci_data.append({'Confidence Level': 'Std Dev', 'Lower Bound': '', 'Upper Bound': f"${cost_ci['std_error']:,.0f}", 'Range': ''})
            st.dataframe(pd.DataFrame(cost_ci_data), use_container_width=True, hide_index=True)
    
    st.markdown('<div class="card">', unsafe_allow_html=True)
    st.markdown('<div class="card-title">📈 Simulation Convergence Analysis</div>', unsafe_allow_html=True)
    
    variation, p50_history = check_convergence(durations, window=1000)
    if variation is not None:
        if variation < 0.01:
            st.success(f"✅ Converged successfully. Variation: {variation:.4%}")
        elif variation < 0.03:
            st.info(f"⚠️ Moderate convergence. Variation: {variation:.4%}")
        else:
            st.warning(f"❌ High variation ({variation:.4%}). Consider increasing iterations.")
    else:
        st.info(f"Running {iterations} iterations. For convergence analysis, 2000+ iterations recommended.")
    
    if p50_history and len(p50_history) > 1:
        fig_conv = go.Figure()
        fig_conv.add_trace(go.Scatter(
            x=list(range(1000, len(durations), 1000)),
            y=p50_history,
            mode='lines+markers',
            name='P50 Value',
            line=dict(color='#3b82f6', width=2),
            marker=dict(size=6, color='#60a5fa')
        ))
        fig_conv.update_layout(
            title="P50 Convergence Over Iterations",
            xaxis_title="Iterations Completed",
            yaxis_title="P50 Duration (days)",
            height=400,
            paper_bgcolor='#1e293b',
            plot_bgcolor='#1e293b',
            font_color='#cbd5e1'
        )
        st.plotly_chart(fig_conv, use_container_width=True, config=chart_config)
    
    st.markdown('</div>', unsafe_allow_html=True)
    
    st.markdown('<div class="card">', unsafe_allow_html=True)
    st.markdown('<div class="card-title">🔄 Scenario Comparison</div>', unsafe_allow_html=True)
    
    if st.button("Compare Scenarios (Optimistic vs Pessimistic)", use_container_width=True):
        with st.spinner("Running scenario comparisons..."):
            scenarios = {"Optimistic (70%)": 0.7, "Most Likely (100%)": 1.0, "Pessimistic (130%)": 1.3}
            scenario_results = {}
            for name, factor in scenarios.items():
                durations_sc, costs_sc, _, _ = run_monte_carlo_full_optimized(
                    st.session_state.risk_factors, 
                    project_duration * factor, 
                    total_cost * factor,
                    indirect_rate, 
                    min(iterations, 5000), 
                    None,
                    None,
                    st.session_state.pra_data
                )
                if durations_sc is not None:
                    scenario_results[name] = {'p80': np.percentile(durations_sc, 80), 'cost_p80': np.percentile(costs_sc, 80)}
            st.session_state.scenario_results = scenario_results
    
    if st.session_state.scenario_results is not None:
        scenario_df = pd.DataFrame([
            {'Scenario': name, 'P80 Duration (days)': f"{res['p80']:.0f}", 'P80 Cost': f"${res['cost_p80']:,.0f}"}
            for name, res in st.session_state.scenario_results.items()
        ])
        st.dataframe(scenario_df, use_container_width=True, hide_index=True)
        
        fig_scenario = go.Figure()
        for name, res in st.session_state.scenario_results.items():
            color = '#22c55e' if 'Optimistic' in name else '#eab308' if 'Most Likely' in name else '#ef4444'
            fig_scenario.add_trace(go.Bar(x=[name], y=[res['p80']], name=name, marker_color=color, text=f"{res['p80']:.0f}", textposition='outside'))
        fig_scenario.update_layout(
            title="Scenario Comparison - P80 Duration",
            xaxis_title="Scenario",
            yaxis_title="Duration (days)",
            height=450,
            paper_bgcolor='#1e293b',
            plot_bgcolor='#1e293b',
            font_color='#cbd5e1'
        )
        st.plotly_chart(fig_scenario, use_container_width=True, config=chart_config)
    
    st.markdown('</div>', unsafe_allow_html=True)
    
    # Project Duration S-Curve
    st.markdown('<div class="card">', unsafe_allow_html=True)
    st.markdown('<div class="card-title">📈 Project Duration S-Curve</div>', unsafe_allow_html=True)
    
    sorted_indices = np.argsort(durations)
    sorted_durations = durations[sorted_indices]
    cumulative = np.arange(1, len(sorted_durations) + 1) / (len(sorted_durations) + 1)
    extended_durations = np.insert(sorted_durations, 0, sorted_durations[0] * 0.95)
    extended_cumulative = np.insert(cumulative, 0, 0)
    
    fig_scurve = go.Figure()
    fig_scurve.add_trace(go.Scatter(
        x=extended_durations, 
        y=extended_cumulative, 
        mode='lines', 
        name='Cumulative Probability',
        line=dict(color='#3b82f6', width=3, shape='spline'),
        hovertemplate='Duration: %{x:.0f} days<br>Probability: %{y:.1%}<extra></extra>'
    ))
    
    for idx, (name, value) in enumerate(percentiles.items()):
        prob = int(name.replace("P", "")) / 100
        fig_scurve.add_hline(
            y=prob,
            line_color=percentile_colors[idx % len(percentile_colors)],
            line_dash="dash",
            line_width=2,
            name=f"{name}: {value:.0f} days",
            legendgroup=name,
            showlegend=True
        )
    
    fig_scurve.update_layout(
        title="<b>Cumulative Probability Curve (S-Curve)</b> - Percentiles shown in legend",
        xaxis_title="<b>Days</b>",
        yaxis_title="<b>Cumulative Probability</b>",
        height=500,
        paper_bgcolor='#1e293b',
        plot_bgcolor='#1e293b',
        font_color='#cbd5e1',
        font=dict(size=12),
        xaxis=dict(fixedrange=True, range=[extended_durations[0], extended_durations[-1] * 1.02], 
                   title_font=dict(size=13, weight='bold')),
        yaxis=dict(fixedrange=True, range=[-0.02, 1.02], title_font=dict(size=13, weight='bold'), tickformat='.0%'),
        legend=dict(yanchor="bottom", y=0.02, xanchor="right", x=0.98, 
                    bgcolor='rgba(30, 41, 59, 0.9)', bordercolor='#334155', borderwidth=1, font=dict(size=10))
    )
    st.plotly_chart(fig_scurve, use_container_width=True, config=chart_config)
    
    st.markdown('</div>', unsafe_allow_html=True)
    
    # Project Duration Distribution
    st.markdown('<div class="card">', unsafe_allow_html=True)
    st.markdown('<div class="card-title">📊 Project Duration Distribution</div>', unsafe_allow_html=True)
    
    hist_counts_dur, bin_edges_dur = np.histogram(durations, bins=50)
    bin_centers_dur = (bin_edges_dur[:-1] + bin_edges_dur[1:]) / 2
    
    fig_dur_hist = go.Figure()
    fig_dur_hist.add_trace(go.Bar(
        x=bin_centers_dur, y=hist_counts_dur, name="Iterations",
        marker_color='#3b82f6', opacity=0.7,
        width=(bin_edges_dur[1] - bin_edges_dur[0]) * 0.85,
        hovertemplate='Duration: %{x:.0f} days<br>Frequency: %{y}<extra></extra>'
    ))
    
    for idx, (name, value) in enumerate(percentiles.items()):
        fig_dur_hist.add_vline(
            x=value,
            line_color=percentile_colors[idx % len(percentile_colors)],
            line_dash="dash",
            line_width=2,
            name=f"{name}: {value:.0f} days",
            legendgroup=name,
            showlegend=True
        )
    
    fig_dur_hist.update_layout(
        title="<b>Project Duration Distribution</b> - Percentiles shown in legend",
        xaxis_title="<b>Days</b>",
        yaxis_title="<b>Frequency</b>",
        height=500,
        paper_bgcolor='#1e293b',
        plot_bgcolor='#1e293b',
        font_color='#cbd5e1',
        font=dict(size=12),
        bargap=0.15,
        xaxis=dict(fixedrange=True, title_font=dict(size=13, weight='bold')),
        yaxis=dict(fixedrange=True, title_font=dict(size=13, weight='bold')),
        legend=dict(yanchor="bottom", y=0.02, xanchor="right", x=0.98, 
                    bgcolor='rgba(30, 41, 59, 0.9)', bordercolor='#334155', borderwidth=1, font=dict(size=10))
    )
    st.plotly_chart(fig_dur_hist, use_container_width=True, config=chart_config)
    
    st.markdown('</div>', unsafe_allow_html=True)
    
    # Project Cost Distribution
    st.markdown('<div class="card">', unsafe_allow_html=True)
    st.markdown('<div class="card-title">💰 Project Cost Distribution</div>', unsafe_allow_html=True)
    
    hist_counts_cost, bin_edges_cost = np.histogram(costs, bins=50)
    bin_centers_cost = (bin_edges_cost[:-1] + bin_edges_cost[1:]) / 2
    
    fig_cost_hist = go.Figure()
    fig_cost_hist.add_trace(go.Bar(
        x=bin_centers_cost, y=hist_counts_cost, name="Iterations",
        marker_color='#10b981', opacity=0.7,
        width=(bin_edges_cost[1] - bin_edges_cost[0]) * 0.85,
        hovertemplate='Cost: $%{x:,.0f}<br>Frequency: %{y}<extra></extra>'
    ))
    
    for idx, (name, value) in enumerate(cost_percentiles.items()):
        fig_cost_hist.add_vline(
            x=value,
            line_color=percentile_colors[idx % len(percentile_colors)],
            line_dash="dash",
            line_width=2,
            name=f"{name}: ${value:,.0f}",
            legendgroup=name,
            showlegend=True
        )
    
    fig_cost_hist.update_layout(
        title="<b>Project Cost Distribution</b> - Percentiles shown in legend",
        xaxis_title="<b>Cost</b>",
        yaxis_title="<b>Frequency</b>",
        height=500,
        paper_bgcolor='#1e293b',
        plot_bgcolor='#1e293b',
        font_color='#cbd5e1',
        font=dict(size=12),
        bargap=0.15,
        xaxis=dict(fixedrange=True, tickformat="$,.0f", title_font=dict(size=13, weight='bold')),
        yaxis=dict(fixedrange=True, title_font=dict(size=13, weight='bold')),
        legend=dict(yanchor="bottom", y=0.02, xanchor="right", x=0.98, 
                    bgcolor='rgba(30, 41, 59, 0.9)', bordercolor='#334155', borderwidth=1, font=dict(size=10))
    )
    st.plotly_chart(fig_cost_hist, use_container_width=True, config=chart_config)
    
    st.markdown('</div>', unsafe_allow_html=True)
    
    # Sensitivity Analysis
    st.markdown('<div class="card">', unsafe_allow_html=True)
    st.markdown('<div class="card-title">📊 Sensitivity Analysis</div>', unsafe_allow_html=True)
    
    st.markdown("#### Duration Sensitivity")
    sensitivity = []
    for risk_name, contributions in risk_contributions.items():
        if len(contributions) > 0:
            sensitivity.append({'Risk': risk_name, 'Impact (%)': np.mean(np.abs(contributions))})
    sensitivity_df = pd.DataFrame(sensitivity).sort_values('Impact (%)', ascending=True)
    
    fig_tornado_duration = go.Figure()
    fig_tornado_duration.add_trace(go.Bar(
        x=sensitivity_df['Impact (%)'], y=sensitivity_df['Risk'], orientation='h',
        marker_color='#3b82f6', name='Impact on Duration',
        hovertemplate='Risk: %{y}<br>Impact on Duration: %{x:.1f}%<extra></extra>',
        text=sensitivity_df['Impact (%)'].apply(lambda x: f'{x:.1f}%'),
        textposition='outside'
    ))
    
    fig_tornado_duration.update_layout(
        title="<b>Risk Impact on Duration (Tornado Chart)</b>",
        xaxis_title="<b>Impact on Duration (%)</b>",
        yaxis_title="<b>Risk Factor</b>",
        height=max(500, len(sensitivity_df) * 30),
        paper_bgcolor='#1e293b',
        plot_bgcolor='#1e293b',
        font_color='#cbd5e1',
        font=dict(size=12),
        xaxis=dict(fixedrange=True, title_font=dict(size=13, weight='bold'), 
                   range=[0, sensitivity_df['Impact (%)'].max() * 1.1]),
        yaxis=dict(fixedrange=True, title_font=dict(size=13, weight='bold'), automargin=True),
        margin=dict(l=150, r=50, t=80, b=50)
    )
    st.plotly_chart(fig_tornado_duration, use_container_width=True, config=chart_config)
    
    st.markdown("#### Cost Sensitivity")
    cost_sensitivity = []
    for risk_name, contributions in risk_contributions.items():
        if len(contributions) > 0:
            cost_impact = np.mean(np.abs(contributions)) * direct_cost * 0.01
            cost_sensitivity.append({'Risk': risk_name, 'Cost Impact ($)': cost_impact})
    cost_sensitivity_df = pd.DataFrame(cost_sensitivity).sort_values('Cost Impact ($)', ascending=True)
    
    if len(cost_sensitivity_df) > 0:
        fig_tornado_cost = go.Figure()
        fig_tornado_cost.add_trace(go.Bar(
            x=cost_sensitivity_df['Cost Impact ($)'],
            y=cost_sensitivity_df['Risk'],
            orientation='h',
            marker_color='#f97316',
            name='Cost Impact',
            hovertemplate='Risk: %{y}<br>Cost Impact: $%{x:,.0f}<extra></extra>',
            text=cost_sensitivity_df['Cost Impact ($)'].apply(lambda x: f'${x:,.0f}'),
            textposition='outside'
        ))
        
        max_cost_impact = cost_sensitivity_df['Cost Impact ($)'].max() * 1.1 if not cost_sensitivity_df.empty else 1
        
        fig_tornado_cost.update_layout(
            title="<b>Risk Impact on Cost (Tornado Chart)</b>",
            xaxis_title="<b>Cost Impact ($)</b>",
            yaxis_title="<b>Risk Factor</b>",
            height=max(500, len(cost_sensitivity_df) * 30),
            paper_bgcolor='#1e293b',
            plot_bgcolor='#1e293b',
            font_color='#cbd5e1',
            font=dict(size=12),
            xaxis=dict(fixedrange=True, title_font=dict(size=13, weight='bold'),
                       range=[0, max_cost_impact], tickformat='$,.0f'),
            yaxis=dict(fixedrange=True, title_font=dict(size=13, weight='bold'), automargin=True),
            margin=dict(l=150, r=50, t=80, b=50)
        )
        st.plotly_chart(fig_tornado_cost, use_container_width=True, config=chart_config)
    else:
        st.info("Run simulation to generate cost sensitivity analysis.")
    
    st.markdown('</div>', unsafe_allow_html=True)
    
    # Risk Contribution Analysis
    st.markdown('<div class="card">', unsafe_allow_html=True)
    st.markdown('<div class="card-title">🥧 Risk Contribution Analysis</div>', unsafe_allow_html=True)
    
    risk_df_pie = pd.DataFrame(sensitivity).sort_values('Impact (%)', ascending=False).head(10)
    fig_pie = px.pie(
        risk_df_pie, values='Impact (%)', names='Risk',
        title="<b>Top 10 Risk Contributors</b> (Hover for details)",
        color_discrete_sequence=px.colors.sequential.Blues_r,
        hole=0.3,
        hover_data={'Impact (%)': ':.1f%'}
    )
    fig_pie.update_layout(
        paper_bgcolor='#1e293b', 
        font_color='#cbd5e1', 
        height=450,
        title_font=dict(size=14, weight='bold')
    )
    st.plotly_chart(fig_pie, use_container_width=True, config=chart_config)
    
    st.markdown('</div>', unsafe_allow_html=True)

# ============================================================================
# CONTINGENCY RESERVE CALCULATION
# ============================================================================
st.markdown('<div class="card">', unsafe_allow_html=True)
st.markdown('<div class="card-title">🎯 CONTINGENCY RESERVE CALCULATION</div>', unsafe_allow_html=True)

col1, col2 = st.columns(2)
with col1:
    confidence_percentile = st.number_input("Confidence Level (%)", min_value=1, max_value=99, value=80, step=5)
with col2:
    contingency_method = st.selectbox("Calculation Method", ["From Monte Carlo", "From EMV", "P80 - P50", "P90 - P50", "Manual Input"])

if contingency_method == "From Monte Carlo" and st.session_state.simulation_results:
    durations = st.session_state.simulation_results['durations']
    p50 = np.percentile(durations, 50)
    selected_percentile = np.percentile(durations, confidence_percentile)
    contingency = selected_percentile - p50
    st.metric(f"Schedule Contingency at P{confidence_percentile}", f"{contingency:.0f} days")
    st.caption(f"P50 = {p50:.0f} days | P{confidence_percentile} = {selected_percentile:.0f} days")
    
    costs = st.session_state.simulation_results['costs']
    cost_p50 = np.percentile(costs, 50)
    cost_selected = np.percentile(costs, confidence_percentile)
    cost_contingency = cost_selected - cost_p50
    st.metric(f"Cost Contingency at P{confidence_percentile}", f"${cost_contingency:,.0f}")

elif contingency_method == "From EMV":
    total_emv_threat = 0
    total_emv_opportunity = 0
    for _, row in st.session_state.risk_factors.iterrows():
        emv, _ = calculate_correct_emv(row, direct_cost, indirect_rate, project_duration)
        if row['Type'] == 'Threat':
            total_emv_threat += emv
        else:
            total_emv_opportunity += emv
    net_emv = total_emv_opportunity - total_emv_threat
    contingency_days = abs(net_emv / indirect_rate) if indirect_rate > 0 else 0
    st.metric("Schedule Contingency from EMV", f"{contingency_days:.0f} days")
    st.metric("Cost Contingency from EMV", f"${abs(net_emv):,.0f}")

elif contingency_method in ["P80 - P50", "P90 - P50"] and st.session_state.simulation_results:
    durations = st.session_state.simulation_results['durations']
    p50 = np.percentile(durations, 50)
    target_percentile = 80 if contingency_method == "P80 - P50" else 90
    target_value = np.percentile(durations, target_percentile)
    contingency = target_value - p50
    st.metric(f"Schedule Contingency ({contingency_method})", f"{contingency:.0f} days")
    st.caption(f"P50 = {p50:.0f} days | {contingency_method} = {target_value:.0f} days")
    
    costs = st.session_state.simulation_results['costs']
    cost_p50 = np.percentile(costs, 50)
    cost_target = np.percentile(costs, target_percentile)
    cost_contingency = cost_target - cost_p50
    st.metric(f"Cost Contingency ({contingency_method})", f"${cost_contingency:,.0f}")

else:
    manual_contingency = st.number_input("Manual Contingency (days)", min_value=0, value=50, step=10)
    st.metric("Manual Schedule Contingency", f"{manual_contingency:.0f} days")

st.markdown('</div>', unsafe_allow_html=True)

# ============================================================================
# POWER BI INTEGRATION
# ============================================================================
st.markdown('<div class="card">', unsafe_allow_html=True)
st.markdown('<div class="card-title">📊 POWER BI INTEGRATION</div>', unsafe_allow_html=True)

st.markdown("""
**Power BI Integration Guide:**
1. Export your simulation data to CSV.
2. Download the Power BI Template Excel file.
3. Open the Excel file in Power BI (Get Data → Excel).
4. Power BI will auto-detect tables and relationships.
5. Create your dashboard visuals using the provided tables.
""")

col_pb1, col_pb2 = st.columns(2)

with col_pb1:
    if st.button("📈 EXPORT SIMULATION DATA (CSV)", use_container_width=True):
        if st.session_state.simulation_details is not None:
            powerbi_df = st.session_state.simulation_details.copy()
            output = io.BytesIO()
            powerbi_df.to_csv(output, index=False)
            output.seek(0)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            st.download_button(label="DOWNLOAD CSV", data=output, file_name=f"powerbi_risk_data_{timestamp}.csv", mime="text/csv", key="powerbi_export")
            st.success("CSV exported successfully!")
        else:
            st.warning("Run simulation first to export data")

with col_pb2:
    if st.button("🔌 GENERATE POWER BI TEMPLATE (EXCEL)", use_container_width=True):
        try:
            powerbi_buffer = io.BytesIO()
            with pd.ExcelWriter(powerbi_buffer, engine='openpyxl') as writer:
                if st.session_state.simulation_details is not None:
                    st.session_state.simulation_details.to_excel(writer, sheet_name='Simulation_Data', index=False)
                else:
                    placeholder_df = pd.DataFrame({
                        'Iteration': [1, 2], 'Duration (days)': [project_duration, project_duration], 
                        'Cost': [direct_cost + indirect_cost, direct_cost + indirect_cost]
                    })
                    placeholder_df.to_excel(writer, sheet_name='Simulation_Data', index=False)
                
                st.session_state.risk_factors.to_excel(writer, sheet_name='Risk_Register', index=False)
                
                risk_names = st.session_state.risk_factors['Risk Name'].tolist()
                corr_matrix_export = make_positive_definite(corr_matrix)
                corr_df = pd.DataFrame(corr_matrix_export, index=risk_names, columns=risk_names)
                corr_df.to_excel(writer, sheet_name='Correlation_Matrix')
                
                metadata = pd.DataFrame({
                    'Parameter': ['Project Name', 'Duration', 'Direct Cost', 'Indirect Rate', 'Generated On'],
                    'Value': [project_name, project_duration, direct_cost, indirect_rate, datetime.now().strftime('%Y-%m-%d %H:%M:%S')]
                })
                metadata.to_excel(writer, sheet_name='Metadata', index=False)
                
                readme_data = pd.DataFrame({
                    'Table': ['Simulation_Data', 'Risk_Register', 'Correlation_Matrix', 'Metadata'],
                    'Description': [
                        'Monte Carlo simulation results (each row is one iteration)',
                        'Risk register with probabilities, impacts, and responses',
                        'Risk correlation matrix for sensitivity analysis',
                        'Project metadata and parameters'
                    ],
                    'Key Columns for Dashboard': [
                        'Duration (days), Cost',
                        'Risk Name, Probability, Impact Value',
                        'All columns (correlation coefficients)',
                        'Project Name, Duration, Direct Cost'
                    ]
                })
                readme_data.to_excel(writer, sheet_name='README_PowerBI', index=False)
            
            powerbi_buffer.seek(0)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            st.download_button(
                label="DOWNLOAD POWER BI TEMPLATE",
                data=powerbi_buffer,
                file_name=f"powerbi_template_{timestamp}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="powerbi_template"
            )
            st.success("Power BI template created! Load this file in Power BI.")
        except Exception as e:
            st.error(f"Error creating template: {e}")

st.markdown("**Power BI Dashboard Visuals to Create:**")
st.markdown("""
- **Duration Distribution**: Histogram using 'Duration (days)' from Simulation_Data
- **Cost Distribution**: Histogram using 'Cost' from Simulation_Data
- **Risk Matrix**: Matrix visual using 'Probability (0-1)' and 'Impact (0-1)' from Risk_Register
- **S-Curve**: Line chart with cumulative probability
- **Tornado Chart**: Bar chart of 'Impact (0-1)' by 'Risk Name'
""")

st.markdown('</div>', unsafe_allow_html=True)

# ============================================================================
# EXPORT & REPORTING
# ============================================================================
st.markdown('<div class="card">', unsafe_allow_html=True)
st.markdown('<div class="card-title">📁 EXPORT & REPORTING</div>', unsafe_allow_html=True)

col_exp1, col_exp2, col_exp3 = st.columns(3)

with col_exp1:
    if st.button("📊 EXPORT RISK REGISTER (EXCEL)", use_container_width=True):
        output = create_professional_excel(st.session_state.risk_factors, 'Risk Register', 'RISK REGISTER')
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        st.download_button(label="DOWNLOAD EXCEL", data=output, file_name=f"risk_register_{timestamp}.xlsx", 
                          mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key="risk_export")

with col_exp2:
    if st.button("📈 EXPORT SIMULATION DETAILS (EXCEL)", use_container_width=True):
        if st.session_state.simulation_details is not None:
            output = create_professional_excel(st.session_state.simulation_details, 'Simulation Details', 'MONTE CARLO SIMULATION RESULTS')
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            st.download_button(label="DOWNLOAD EXCEL", data=output, file_name=f"simulation_details_{timestamp}.xlsx",
                              mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key="sim_details_export")
        else:
            st.warning("Run simulation first")

with col_exp3:
    if st.button("📑 EXPORT ALL CHARTS (ZIP)", use_container_width=True):
        if st.session_state.simulation_results is not None:
            with st.spinner("Generating charts ZIP file..."):
                current_durations = st.session_state.simulation_results['durations']
                current_costs = st.session_state.simulation_results['costs']
                current_risk_contributions = st.session_state.simulation_results['risk_contributions']
                
                export_percentiles = {}
                for p in user_percentiles:
                    if 0 <= p <= 100:
                        export_percentiles[f'P{p}'] = np.percentile(current_durations, p)
                
                export_cost_percentiles = {}
                for p in user_percentiles:
                    if 0 <= p <= 100:
                        export_cost_percentiles[f'P{p}'] = np.percentile(current_costs, p)
                
                zip_buffer = create_zip_with_charts(
                    current_durations,
                    current_costs,
                    current_risk_contributions,
                    user_percentiles,
                    export_percentiles,
                    export_cost_percentiles
                )
                if zip_buffer:
                    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                    st.download_button(label="DOWNLOAD ZIP", data=zip_buffer, file_name=f"charts_{timestamp}.zip", 
                                      mime="application/zip", key="charts_export")
        else:
            st.warning("Run simulation first")

st.markdown('</div>', unsafe_allow_html=True)

# ============================================================================
# COMPLETE SAVE / LOAD / RESET
# ============================================================================
st.markdown('<div class="card">', unsafe_allow_html=True)
st.markdown('<div class="card-title">💾 COMPLETE SAVE / LOAD / RESET</div>', unsafe_allow_html=True)

col_save, col_reset = st.columns(2)

with col_save:
    if st.button("💾 SAVE COMPLETE PROJECT", use_container_width=True):
        try:
            complete_project = {
                'version': '2.0',
                'timestamp': datetime.now().isoformat(),
                'risk_factors': st.session_state.risk_factors.to_dict('records'),
                'risk_thresholds': st.session_state.risk_thresholds,
                'project_parameters': {
                    'project_name': project_name,
                    'project_location': project_location,
                    'project_manager': project_manager,
                    'project_duration': project_duration,
                    'direct_cost': direct_cost,
                    'indirect_rate': indirect_rate,
                    'detected_currency': detected_currency
                },
                'simulation_settings': {
                    'iterations': iterations,
                    'percentiles': percentile_input
                },
                'decision_tree_data': {
                    'decision_type': decision_type,
                    'decision_name': decision_name,
                    'num_alternatives': num_alternatives,
                    'alternatives': alternatives
                }
            }
            
            json_data = json.dumps(complete_project, indent=2, default=str)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            
            st.download_button(
                label="DOWNLOAD COMPLETE PROJECT (.json)",
                data=json_data,
                file_name=f"complete_project_{timestamp}.json",
                mime="application/json",
                key="save_complete_btn"
            )
            st.toast("✅ Complete project ready for download!", icon="💾")
        except Exception as e:
            st.error(f"Error saving: {e}")

with col_reset:
    if st.button("🔄 RESET TO DEFAULTS", use_container_width=True, type="secondary"):
        default_risk_factors = pd.DataFrame({
            'Risk Name': ['Geopolitical Instability', 'Pandemic Disruption', 'Supply Chain Disruption', 'Labor Shortage', 'Weather Events'],
            'Type': ['Threat', 'Threat', 'Threat', 'Threat', 'Threat'],
            'Probability (0-1)': [0.40, 0.35, 0.30, 0.25, 0.15],
            'Impact (0-1)': [0.25, 0.20, 0.15, 0.12, 0.08],
            'Distribution': ['Triangular', 'Triangular', 'Triangular', 'Triangular', 'Triangular'],
            'Optimistic (0-1)': [0.10, 0.05, 0.05, 0.05, 0.02],
            'Most Likely (0-1)': [0.25, 0.20, 0.15, 0.12, 0.08],
            'Pessimistic (0-1)': [0.40, 0.35, 0.30, 0.25, 0.15],
            'Correlated With': ['', 'Geopolitical Instability', 'Geopolitical Instability', '', ''],
            'Correlation Strength': [0.0, 0.8, 0.7, 0.0, 0.0],
            'Response Strategy': ['Transfer', 'Mitigate', 'Mitigate', 'Mitigate', 'Accept'],
            'Response Plan': ['Political risk insurance, diversify suppliers', 'Health protocols, remote work plans', 'Dual sourcing, inventory buffer', 'Recruitment plan, training', 'Weather monitoring, schedule flexibility']
        })
        
        st.session_state.risk_factors = default_risk_factors
        st.session_state.risk_thresholds = {
            'high_prob': 0.7, 'medium_prob': 0.5, 'low_prob': 0.3,
            'high_impact': 0.7, 'medium_impact': 0.5, 'low_impact': 0.3
        }
        st.session_state.ai_trained = False
        st.session_state.ai_model = None
        st.session_state.simulation_results = None
        st.session_state.simulation_run_flag = False
        st.session_state.simulation_details = None
        st.session_state.scenario_results = None
        
        if 'uploaded_file_name' in st.session_state:
            del st.session_state.uploaded_file_name
        if 'last_loaded_file_key' in st.session_state:
            del st.session_state.last_loaded_file_key
        if 'uploaded_project_name' in st.session_state:
            del st.session_state.uploaded_project_name
        if 'last_loaded_project_key' in st.session_state:
            del st.session_state.last_loaded_project_key
        if 'geocode_cache' in st.session_state:
            del st.session_state.geocode_cache
        if 'p6_wbs_summary' in st.session_state:
            del st.session_state.p6_wbs_summary
        
        st.session_state.p6_activities = None
        st.session_state.file_uploader_key_counter += 1
        
        st.toast("✅ Reset to default! All data cleared.", icon="🔄")
        time.sleep(0.5)
        st.rerun()

st.markdown("---")

uploader_key = f"load_complete_project_{st.session_state.file_uploader_key_counter}"
uploaded_project = st.file_uploader("📂 LOAD COMPLETE PROJECT (.json) - auto-loads when selected", type=["json"], key=uploader_key)

if uploaded_project is not None:
    current_file_key = f"{uploaded_project.name}_{uploaded_project.size}"
    
    if st.session_state.get('last_loaded_project_key') != current_file_key:
        try:
            project_data = json.load(uploaded_project)
            
            if 'risk_factors' in project_data:
                st.session_state.risk_factors = pd.DataFrame(project_data['risk_factors'])
            
            if 'risk_thresholds' in project_data:
                st.session_state.risk_thresholds = project_data['risk_thresholds']
            
            if 'project_parameters' in project_data:
                st.session_state.loaded_project_params = project_data['project_parameters']
            
            if 'simulation_settings' in project_data:
                st.session_state.loaded_simulation_settings = project_data['simulation_settings']
            
            if 'decision_tree_data' in project_data:
                st.session_state.loaded_decision_tree = project_data['decision_tree_data']
            
            st.session_state.ai_trained = False
            st.session_state.ai_model = None
            st.session_state.simulation_results = None
            st.session_state.simulation_run_flag = False
            st.session_state.simulation_details = None
            
            st.session_state.uploaded_project_name = uploaded_project.name
            st.session_state.last_loaded_project_key = current_file_key
            
            st.toast(f"✅ Complete project loaded from {uploaded_project.name}!", icon="✅")
            time.sleep(0.5)
            st.rerun()
            
        except Exception as e:
            st.error(f"Error loading project: {e}")

if st.session_state.get('uploaded_project_name'):
    st.info(f"📁 Loaded project: **{st.session_state.uploaded_project_name}**")
    
    with st.expander("📋 Loaded Project Details"):
        if st.session_state.get('loaded_project_params'):
            params = st.session_state.loaded_project_params
            st.markdown(f"**Project Name:** {params.get('project_name', 'N/A')}")
            st.markdown(f"**Location:** {params.get('project_location', 'N/A')}")
            st.markdown(f"**Duration:** {params.get('project_duration', 'N/A')} days")
            st.markdown(f"**Direct Cost:** ${params.get('direct_cost', 0):,.0f}")
        
        if st.session_state.get('loaded_simulation_settings'):
            settings = st.session_state.loaded_simulation_settings
            st.markdown(f"**Simulation Iterations:** {settings.get('iterations', 'N/A')}")
            st.markdown(f"**Percentiles:** {settings.get('percentiles', 'N/A')}")
        
        if st.session_state.get('loaded_decision_tree'):
            dt = st.session_state.loaded_decision_tree
            st.markdown(f"**Decision Type:** {dt.get('decision_type', 'N/A')}")
            st.markdown(f"**Number of Alternatives:** {dt.get('num_alternatives', 'N/A')}")

st.markdown('</div>', unsafe_allow_html=True)

# ============================================================================
# FOOTER
# ============================================================================
st.markdown("""
<div class="footer">
    RISK ANALYSIS PLATFORM | ENTERPRISE EDITION | PMI-RMP COMPLIANT | VERSION 2.0<br>
    Features: Monte Carlo Simulation | Correlation Matrix | AI Risk Prediction | EMV Analysis<br>
    Contingency Calculation | Sensitivity Analysis | Scenario Comparison | Convergence Check<br>
    Response Plans | Real-Time Weather | Primavera P6 Integration | Power BI Export<br>
    Professional Excel Export | Chart Export | Complete Project Save/Load
</div>
""", unsafe_allow_html=True)
